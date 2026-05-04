from __future__ import annotations

import os
import shutil
import sqlite3
import psycopg2
import requests
from psycopg2.extras import RealDictCursor
from urllib.parse import urlparse
from datetime import datetime, date, timedelta
from functools import wraps
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

from flask import (
    Flask,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
DATABASE = INSTANCE_DIR / "relogio_ponto.db"
EXPORTS_DIR = BASE_DIR / "exports"
BACKUPS_DIR = BASE_DIR / "backups"

WEEKDAY_LABELS = {
    0: "Segunda-feira",
    1: "Terça-feira",
    2: "Quarta-feira",
    3: "Quinta-feira",
    4: "Sexta-feira",
    5: "Sábado",
    6: "Domingo",
}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "petegato_relogio_2026")
app.config["DATABASE"] = str(DATABASE)
app.config["EXPORTS_DIR"] = str(EXPORTS_DIR)
app.config["BACKUPS_DIR"] = str(BACKUPS_DIR)

STORE_ALLOWED_IP = os.environ.get("STORE_ALLOWED_IP", "").strip()
ADMIN_SECRET_KEY = os.environ.get("ADMIN_SECRET_KEY", "").strip()


def get_client_ip() -> str:
    """Retorna o IP real do usuário atrás do proxy do Render."""
    for header in ("X-Forwarded-For", "X-Real-IP", "CF-Connecting-IP", "True-Client-IP"):
        value = request.headers.get(header, "")
        if value:
            return value.split(",")[0].strip()
    return request.remote_addr or ""


def is_store_network() -> bool:
    client_ip = get_client_ip()
    return bool(STORE_ALLOWED_IP and client_ip == STORE_ALLOWED_IP)


def has_admin_key() -> bool:
    key = request.args.get("chave", "").strip()
    if key and ADMIN_SECRET_KEY and key == ADMIN_SECRET_KEY:
        session["admin_key_ok"] = True
        return True
    return bool(session.get("admin_key_ok"))


@app.before_request
def security_gate():
    """Segurança corrigida para SaaS + Checkout Pro.

    - /ponto e / continuam protegidos por IP da loja ou chave admin.
    - /login, /assinatura, /assinatura/pagar, /webhook, /webhook/mercadopago e /pagamento/* ficam livres do bloqueio por IP/chave.
      A proteção de login continua sendo feita pelo @admin_required nas rotas internas.
    - Painel administrativo de clientes usa sessão/login, não bloqueio por IP.
    """
    allowed_endpoints = {"static", "service_worker", "health"}
    if request.endpoint in allowed_endpoints:
        return None

    public_prefixes = (
        "/login",
        "/assinatura",
        "/assinatura/pagar",
        "/webhook/mercadopago",
        "/pagamento",
    )
    if request.path.startswith(public_prefixes):
        return None

    # Tela de ponto pública da loja: só libera na rede autorizada ou com chave segura.
    if request.path.startswith("/ponto") or request.path == "/":
        if not is_store_network() and not has_admin_key():
            return render_template("blocked.html", ip=get_client_ip()), 403

    return None


# ---------------------------
# Database helpers
# ---------------------------
def get_db():
    if "db" not in g:
        database_url = os.getenv("DATABASE_URL")

        if database_url:
            g.db = psycopg2.connect(database_url, cursor_factory=RealDictCursor)
        else:
            g.db = sqlite3.connect(app.config["DATABASE"])
            g.db.row_factory = sqlite3.Row

    return g.db


@app.teardown_appcontext
def close_db(exc: Optional[BaseException]) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


# ---------------------------
# Utils
# ---------------------------
def now_dt() -> datetime:
    return datetime.now()


def now_iso() -> str:
    return now_dt().strftime("%Y-%m-%d %H:%M:%S")


def today_str() -> str:
    return date.today().strftime("%Y-%m-%d")


def parse_dt(value: str) -> datetime:
    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_hhmm(value: str) -> Tuple[int, int]:
    if not value or ":" not in str(value):
        raise ValueError(f"Horário inválido ou vazio: {value!r}")
    hh, mm = str(value).split(":")[:2]
    return int(hh), int(mm)


def combine_day_time(day: date, hhmm: str) -> datetime:
    hh, mm = parse_hhmm(hhmm)
    return datetime(day.year, day.month, day.day, hh, mm)


def format_dt(value: Optional[str]) -> str:
    if not value:
        return "-"
    return parse_dt(value).strftime("%d/%m/%Y %H:%M:%S")


def format_minutes(total_minutes: int) -> str:
    sign = "-" if total_minutes < 0 else ""
    total_minutes = abs(total_minutes)
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{sign}{hours:02d}:{minutes:02d}"


def record_sequence() -> List[str]:
    return ["ENTRADA", "SAIDA_INTERVALO", "RETORNO_INTERVALO", "SAIDA"]


def normalize_phone(phone: str) -> str:
    return "".join(ch for ch in phone if ch.isdigit())


def build_whatsapp_link(phone: str, message: str) -> Optional[str]:
    digits = normalize_phone(phone)
    if not digits:
        return None
    return f"https://wa.me/{digits}?text={quote(message)}"


def weekday_name(idx: int) -> str:
    return WEEKDAY_LABELS.get(idx, str(idx))


def adapt_query(query: str) -> str:
    if os.getenv("DATABASE_URL"):
        return query.replace("?", "%s")
    return query


def query_db(query: str, params: tuple = (), one: bool = False) -> Any:
    cur = get_db().cursor()
    cur.execute(adapt_query(query), params)
    rows = cur.fetchall()
    cur.close()
    return (rows[0] if rows else None) if one else rows


def execute_db(query: str, params: tuple = ()) -> int:
    db = get_db()
    cur = db.cursor()
    is_postgres = bool(os.getenv("DATABASE_URL"))
    sql = adapt_query(query)
    lastrowid = 0

    try:
        if is_postgres:
            normalized = sql.strip().lower()
            needs_returning = (
                normalized.startswith("insert into")
                and " returning " not in normalized
                and not normalized.startswith("insert into settings")
            )
            if needs_returning:
                sql_to_run = sql.rstrip().rstrip(";") + " RETURNING id"
                cur.execute(sql_to_run, params)
                row = cur.fetchone()
                if row:
                    lastrowid = row["id"] if isinstance(row, dict) else row[0]
            else:
                cur.execute(sql, params)
        else:
            cur.execute(sql, params)
            lastrowid = cur.lastrowid

        db.commit()
        return lastrowid
    except Exception:
        db.rollback()
        raise
    finally:
        cur.close()


def get_setting(key: str, default: str = "") -> str:
    row = query_db("SELECT value FROM settings WHERE key = ?", (key,), one=True)
    return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    execute_db(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
        """,
        (key, value, now_iso()),
    )


def init_db() -> None:
    INSTANCE_DIR.mkdir(parents=True, exist_ok=True)
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)

    database_url = os.getenv("DATABASE_URL")

    if database_url:
        with psycopg2.connect(database_url, cursor_factory=RealDictCursor) as conn:
            with conn.cursor() as cur:
                cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'admin',
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS schedules (
                    id SERIAL PRIMARY KEY,
                    description TEXT NOT NULL,
                    entry_time TEXT NOT NULL,
                    lunch_out_time TEXT NOT NULL,
                    lunch_in_time TEXT NOT NULL,
                    exit_time TEXT NOT NULL,
                    tolerance_minutes INTEGER NOT NULL DEFAULT 10,
                    active INTEGER NOT NULL DEFAULT 1
                );

                CREATE TABLE IF NOT EXISTS employees (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    cpf TEXT,
                    registration TEXT NOT NULL UNIQUE,
                    role_name TEXT,
                    phone TEXT,
                    admission_date TEXT,
                    pin_hash TEXT NOT NULL,
                    schedule_id INTEGER REFERENCES schedules(id),
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS employee_schedule_days (
                    id SERIAL PRIMARY KEY,
                    employee_id INTEGER NOT NULL REFERENCES employees(id),
                    weekday INTEGER NOT NULL,
                    schedule_id INTEGER NOT NULL REFERENCES schedules(id),
                    UNIQUE(employee_id, weekday)
                );

                CREATE TABLE IF NOT EXISTS time_records (
                    id SERIAL PRIMARY KEY,
                    employee_id INTEGER NOT NULL REFERENCES employees(id),
                    record_type TEXT NOT NULL,
                    record_time TEXT NOT NULL,
                    notes TEXT,
                    ip_origin TEXT,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS time_adjustments (
                    id SERIAL PRIMARY KEY,
                    record_id INTEGER REFERENCES time_records(id),
                    employee_id INTEGER NOT NULL REFERENCES employees(id),
                    admin_user_id INTEGER NOT NULL REFERENCES users(id),
                    old_value TEXT,
                    new_value TEXT NOT NULL,
                    reason TEXT NOT NULL,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS system_logs (
                    id SERIAL PRIMARY KEY,
                    user_name TEXT,
                    action TEXT NOT NULL,
                    details TEXT,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                """)

                cur.execute("SELECT id FROM schedules LIMIT 1")
                if not cur.fetchone():
                    cur.execute("""
                        INSERT INTO schedules (
                            description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes, active
                        ) VALUES
                        (%s, %s, %s, %s, %s, %s, %s),
                        (%s, %s, %s, %s, %s, %s, %s),
                        (%s, %s, %s, %s, %s, %s, %s)
                    """, (
                        "Jornada padrão 08h às 17h", "08:00", "12:00", "13:00", "17:00", 10, 1,
                        "Terça a Sexta 09h às 18h", "09:00", "12:00", "13:00", "18:00", 10, 1,
                        "Sábado 08h às 17h", "08:00", "12:00", "13:00", "17:00", 10, 1,
                    ))

                cur.execute("SELECT id FROM users WHERE username = %s", ("admin",))
                if not cur.fetchone():
                    cur.execute("""
                        INSERT INTO users (name, username, password_hash, role, active, created_at)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (
                        "Administrador",
                        "admin",
                        generate_password_hash("Pet&gato3264"),
                        "admin",
                        1,
                        now_iso(),
                    ))

                cur.execute("SELECT key FROM settings WHERE key = %s", ("company_whatsapp",))
                if not cur.fetchone():
                    cur.execute(
                        "INSERT INTO settings (key, value, updated_at) VALUES (%s, %s, %s)",
                        ("company_whatsapp", "", now_iso()),
                    )
            conn.commit()
        return

    schema = """
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'admin',
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS schedules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        description TEXT NOT NULL,
        entry_time TEXT NOT NULL,
        lunch_out_time TEXT NOT NULL,
        lunch_in_time TEXT NOT NULL,
        exit_time TEXT NOT NULL,
        tolerance_minutes INTEGER NOT NULL DEFAULT 10,
        active INTEGER NOT NULL DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        cpf TEXT,
        registration TEXT NOT NULL UNIQUE,
        role_name TEXT,
        phone TEXT,
        admission_date TEXT,
        pin_hash TEXT NOT NULL,
        schedule_id INTEGER,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL,
        FOREIGN KEY (schedule_id) REFERENCES schedules(id)
    );

    CREATE TABLE IF NOT EXISTS employee_schedule_days (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        weekday INTEGER NOT NULL,
        schedule_id INTEGER NOT NULL,
        UNIQUE(employee_id, weekday),
        FOREIGN KEY (employee_id) REFERENCES employees(id),
        FOREIGN KEY (schedule_id) REFERENCES schedules(id)
    );

    CREATE TABLE IF NOT EXISTS time_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        record_type TEXT NOT NULL,
        record_time TEXT NOT NULL,
        notes TEXT,
        ip_origin TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY (employee_id) REFERENCES employees(id)
    );

    CREATE TABLE IF NOT EXISTS time_adjustments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_id INTEGER,
        employee_id INTEGER NOT NULL,
        admin_user_id INTEGER NOT NULL,
        old_value TEXT,
        new_value TEXT NOT NULL,
        reason TEXT NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY (record_id) REFERENCES time_records(id),
        FOREIGN KEY (employee_id) REFERENCES employees(id),
        FOREIGN KEY (admin_user_id) REFERENCES users(id)
    );

    CREATE TABLE IF NOT EXISTS system_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_name TEXT,
        action TEXT NOT NULL,
        details TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL,
        updated_at TEXT NOT NULL
    );
    """

    with sqlite3.connect(app.config["DATABASE"]) as conn:
        conn.executescript(schema)
        conn.commit()

    with sqlite3.connect(app.config["DATABASE"]) as conn:
        conn.row_factory = sqlite3.Row
        schedule = conn.execute("SELECT id FROM schedules LIMIT 1").fetchone()
        if not schedule:
            conn.execute(
                """
                INSERT INTO schedules (
                    description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes, active
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                ("Jornada padrão 08h às 17h", "08:00", "12:00", "13:00", "17:00", 10, 1),
            )
            conn.execute(
                """
                INSERT INTO schedules (
                    description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes, active
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                ("Terça a Sexta 09h às 18h", "09:00", "12:00", "13:00", "18:00", 10, 1),
            )
            conn.execute(
                """
                INSERT INTO schedules (
                    description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes, active
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                ("Sábado 08h às 17h", "08:00", "12:00", "13:00", "17:00", 10, 1),
            )

        admin = conn.execute("SELECT id FROM users WHERE username = ?", ("admin",)).fetchone()
        if not admin:
            conn.execute(
                "INSERT INTO users (name, username, password_hash, role, active, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    "Administrador",
                    "admin",
                    generate_password_hash("Pet&gato3264"),
                    "admin",
                    1,
                    now_iso(),
                ),
            )

        existing = conn.execute("SELECT key FROM settings WHERE key = 'company_whatsapp'").fetchone()
        if not existing:
            conn.execute(
                "INSERT INTO settings (key, value, updated_at) VALUES (?, ?, ?)",
                ("company_whatsapp", "", now_iso()),
            )
        conn.commit()


def get_schedule_for_day(employee_id: int, ref_date: str) -> Optional[sqlite3.Row]:
    ref = parse_date(ref_date)
    weekday = ref.weekday()
    schedule = query_db(
        """
        SELECT s.*
        FROM employee_schedule_days esd
        JOIN schedules s ON s.id = esd.schedule_id
        WHERE esd.employee_id = ? AND esd.weekday = ? AND s.active = 1
        """,
        (employee_id, weekday),
        one=True,
    )
    if schedule:
        return schedule

    return query_db(
        "SELECT s.* FROM employees e LEFT JOIN schedules s ON s.id = e.schedule_id WHERE e.id = ?",
        (employee_id,),
        one=True,
    )


def get_employee_week_schedule_map(employee_id: int) -> Dict[int, sqlite3.Row]:
    rows = query_db(
        """
        SELECT esd.weekday, s.id AS schedule_id, s.description
        FROM employee_schedule_days esd
        JOIN schedules s ON s.id = esd.schedule_id
        WHERE esd.employee_id = ?
        ORDER BY esd.weekday
        """,
        (employee_id,),
    )
    return {row["weekday"]: row for row in rows}


def set_employee_week_schedule(employee_id: int, weekday: int, schedule_id: Optional[int]) -> None:
    execute_db("DELETE FROM employee_schedule_days WHERE employee_id = ? AND weekday = ?", (employee_id, weekday))
    if schedule_id:
        execute_db(
            "INSERT INTO employee_schedule_days (employee_id, weekday, schedule_id) VALUES (?, ?, ?)",
            (employee_id, weekday, schedule_id),
        )


def next_record_type(employee_id: int, ref_date: Optional[str] = None) -> str:
    ref_date = ref_date or today_str()
    records = query_db(
        """
        SELECT record_type FROM time_records
        WHERE employee_id = ? AND date(record_time) = ?
        ORDER BY record_time ASC, id ASC
        """,
        (employee_id, ref_date),
    )
    index = len(records)
    sequence = record_sequence()
    if index >= len(sequence):
        return "COMPLETO"
    return sequence[index]


def calculate_day_summary(employee_id: int, ref_date: str) -> Dict[str, Any]:
    employee = query_db(
        """
        SELECT e.*
        FROM employees e
        WHERE e.id = ?
        """,
        (employee_id,),
        one=True,
    )
    records = query_db(
        """
        SELECT * FROM time_records
        WHERE employee_id = ? AND date(record_time) = ?
        ORDER BY record_time ASC, id ASC
        """,
        (employee_id, ref_date),
    )
    schedule = get_schedule_for_day(employee_id, ref_date)
    sequence = {row["record_type"]: row["record_time"] for row in records}
    ref_day = parse_date(ref_date)

    scheduled_entry = combine_day_time(ref_day, schedule["entry_time"]) if schedule and schedule["entry_time"] else None
    scheduled_lunch_out = combine_day_time(ref_day, schedule["lunch_out_time"]) if schedule and schedule["lunch_out_time"] else None
    scheduled_lunch_in = combine_day_time(ref_day, schedule["lunch_in_time"]) if schedule and schedule["lunch_in_time"] else None
    scheduled_exit = combine_day_time(ref_day, schedule["exit_time"]) if schedule and schedule["exit_time"] else None

    worked_minutes = 0
    if sequence.get("ENTRADA") and sequence.get("SAIDA"):
        start = parse_dt(sequence["ENTRADA"])
        end = parse_dt(sequence["SAIDA"])
        worked_minutes = int((end - start).total_seconds() // 60)
        if sequence.get("SAIDA_INTERVALO") and sequence.get("RETORNO_INTERVALO"):
            lunch_out = parse_dt(sequence["SAIDA_INTERVALO"])
            lunch_in = parse_dt(sequence["RETORNO_INTERVALO"])
            worked_minutes -= int((lunch_in - lunch_out).total_seconds() // 60)

    scheduled_worked = 0
    if scheduled_entry and scheduled_exit and scheduled_lunch_out and scheduled_lunch_in:
        scheduled_worked = int((scheduled_exit - scheduled_entry).total_seconds() // 60)
        scheduled_worked -= int((scheduled_lunch_in - scheduled_lunch_out).total_seconds() // 60)

    delay_minutes = 0
    early_exit_minutes = 0
    extra_minutes = 0
    lunch_less_than_expected = False
    inconsistency = False

    if sequence.get("ENTRADA") and scheduled_entry:
        actual_entry = parse_dt(sequence["ENTRADA"])
        tolerance = schedule["tolerance_minutes"] if schedule else 0
        diff = int((actual_entry - scheduled_entry).total_seconds() // 60)
        if diff > tolerance:
            delay_minutes = diff

    if sequence.get("SAIDA") and scheduled_exit:
        actual_exit = parse_dt(sequence["SAIDA"])
        diff = int((scheduled_exit - actual_exit).total_seconds() // 60)
        if diff > 0:
            early_exit_minutes = diff

    if worked_minutes and scheduled_worked:
        extra_minutes = worked_minutes - scheduled_worked

    if sequence.get("SAIDA_INTERVALO") and sequence.get("RETORNO_INTERVALO") and scheduled_lunch_out and scheduled_lunch_in:
        actual_lunch = int((parse_dt(sequence["RETORNO_INTERVALO"]) - parse_dt(sequence["SAIDA_INTERVALO"])).total_seconds() // 60)
        expected_lunch = int((scheduled_lunch_in - scheduled_lunch_out).total_seconds() // 60)
        lunch_less_than_expected = actual_lunch < expected_lunch

    if len(records) not in (0, 4):
        inconsistency = True

    return {
        "employee": employee,
        "records": records,
        "sequence": sequence,
        "schedule": schedule,
        "worked_minutes": worked_minutes,
        "scheduled_minutes": scheduled_worked,
        "delay_minutes": delay_minutes,
        "early_exit_minutes": early_exit_minutes,
        "extra_minutes": extra_minutes,
        "lunch_less_than_expected": lunch_less_than_expected,
        "inconsistency": inconsistency,
        "next_type": next_record_type(employee_id, ref_date),
        "weekday_name": weekday_name(ref_day.weekday()),
    }



def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            flash("Faça login para acessar o painel.", "warning")
            return redirect(url_for("login"))

        # Super admin sempre entra.
        try:
            if is_super_admin():
                return func(*args, **kwargs)
        except Exception:
            pass

        # Cliente comum precisa estar ativo/pago/teste válido.
        try:
            company = current_company()
            if company and not company_access_allowed(company):
                if request.endpoint != "subscription_page" and not request.path.startswith("/assinatura"):
                    flash("Assinatura pendente ou vencida. Regularize para liberar o painel.", "warning")
                    return redirect(url_for("subscription_page"))
        except Exception as exc:
            print("Aviso verificação assinatura:", exc)

        return func(*args, **kwargs)
    return wrapper


def log_action(action: str, details: str = "") -> None:
    username = session.get("user_name", "Sistema")
    execute_db(
        "INSERT INTO system_logs (user_name, action, details, created_at) VALUES (?, ?, ?, ?)",
        (username, action, details, now_iso()),
    )


def month_bounds(year: int, month: int) -> Tuple[str, str]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def generate_excel_report(year: int, month: int) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Espelho de Ponto"

    headers = [
        "Funcionário", "Matrícula", "Data", "Dia", "Jornada", "Entrada", "Saída Intervalo",
        "Retorno Intervalo", "Saída", "Horas Trabalhadas", "Atraso", "Saída Antecipada",
        "Saldo (Extra/Negativo)", "Inconsistência"
    ]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    start, end = month_bounds(year, month)
    employees = query_db("SELECT * FROM employees WHERE active = 1 ORDER BY name")

    for emp in employees:
        current = parse_date(start)
        end_date = parse_date(end)
        while current < end_date:
            summary = calculate_day_summary(emp["id"], current.strftime("%Y-%m-%d"))
            seq = summary["sequence"]
            ws.append([
                emp["name"],
                emp["registration"],
                current.strftime("%d/%m/%Y"),
                summary["weekday_name"],
                summary["schedule"]["description"] if summary["schedule"] else "Sem jornada",
                parse_dt(seq["ENTRADA"]).strftime("%H:%M:%S") if seq.get("ENTRADA") else "",
                parse_dt(seq["SAIDA_INTERVALO"]).strftime("%H:%M:%S") if seq.get("SAIDA_INTERVALO") else "",
                parse_dt(seq["RETORNO_INTERVALO"]).strftime("%H:%M:%S") if seq.get("RETORNO_INTERVALO") else "",
                parse_dt(seq["SAIDA"]).strftime("%H:%M:%S") if seq.get("SAIDA") else "",
                format_minutes(summary["worked_minutes"]),
                format_minutes(summary["delay_minutes"]),
                format_minutes(summary["early_exit_minutes"]),
                format_minutes(summary["extra_minutes"]),
                "SIM" if summary["inconsistency"] else "NÃO",
            ])
            current += timedelta(days=1)

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[column].width = min(max_length + 2, 30)

    file_path = EXPORTS_DIR / f"relatorio_ponto_{year}_{month:02d}.xlsx"
    wb.save(file_path)
    return file_path


def generate_pdf_report(employee_id: int, year: int, month: int) -> Path:
    employee = query_db("SELECT * FROM employees WHERE id = ?", (employee_id,), one=True)
    file_path = EXPORTS_DIR / f"espelho_ponto_{employee['registration']}_{year}_{month:02d}.pdf"

    doc = SimpleDocTemplate(str(file_path), pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Espelho de Ponto - Pet & Gatô", styles["Title"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Funcionário: {employee['name']} | Matrícula: {employee['registration']}", styles["Normal"]))
    story.append(Paragraph(f"Referência: {month:02d}/{year}", styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [["Data", "Dia", "Jornada", "Entrada", "Saída Intervalo", "Retorno Intervalo", "Saída", "Horas", "Atraso", "Saldo", "Inconsistência"]]
    start, end = month_bounds(year, month)
    current = parse_date(start)
    end_date = parse_date(end)
    while current < end_date:
        summary = calculate_day_summary(employee_id, current.strftime("%Y-%m-%d"))
        seq = summary["sequence"]
        data.append([
            current.strftime("%d/%m/%Y"),
            summary["weekday_name"],
            summary["schedule"]["description"] if summary["schedule"] else "Sem jornada",
            parse_dt(seq["ENTRADA"]).strftime("%H:%M:%S") if seq.get("ENTRADA") else "",
            parse_dt(seq["SAIDA_INTERVALO"]).strftime("%H:%M:%S") if seq.get("SAIDA_INTERVALO") else "",
            parse_dt(seq["RETORNO_INTERVALO"]).strftime("%H:%M:%S") if seq.get("RETORNO_INTERVALO") else "",
            parse_dt(seq["SAIDA"]).strftime("%H:%M:%S") if seq.get("SAIDA") else "",
            format_minutes(summary["worked_minutes"]),
            format_minutes(summary["delay_minutes"]),
            format_minutes(summary["extra_minutes"]),
            "SIM" if summary["inconsistency"] else "NÃO",
        ])
        current += timedelta(days=1)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
    ]))
    story.append(table)
    story.append(Spacer(1, 25))
    story.append(Paragraph("Assinatura do colaborador: ____________________________________________", styles["Normal"]))

    doc.build(story)
    return file_path


def create_backup() -> Path:
    timestamp = now_dt().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUPS_DIR / f"backup_{timestamp}.db"
    if DATABASE.exists():
        shutil.copy2(DATABASE, backup_path)
    return backup_path


# ---------------------------
# Routes - auth
# ---------------------------
@app.route("/")
def index():
    return redirect(url_for("punch"))

def liberar_empresa(slug_empresa):
    conn = get_db()
    conn.execute("""
        UPDATE empresas
        SET status='ATIVO',
            acesso_liberado=1,
            data_pagamento=datetime('now'),
            data_expiracao=datetime('now', '+30 days')
        WHERE slug=?
    """, (slug_empresa,))
    conn.commit()

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = query_db("SELECT * FROM users WHERE username = ? AND active = 1", (username,), one=True)
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["user_name"] = user["name"]
            session["role"] = user.get("role", "admin") if hasattr(user, "get") else user["role"]
            session["company_id"] = user.get("company_id") if hasattr(user, "get") else user["company_id"]
            if session.get("company_id"):
                empresa_login = query_db("SELECT name FROM companies WHERE id = ?", (session["company_id"],), one=True)
                session["company_name"] = empresa_login["name"] if empresa_login else ""
            log_action("LOGIN", f"Usuário {username} entrou no sistema")
            flash("Login realizado com sucesso.", "success")
            return redirect(url_for("dashboard"))
        flash("Usuário ou senha inválidos.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    if session.get("user_id"):
        log_action("LOGOUT", "Usuário saiu do sistema")
    session.clear()
    flash("Sessão encerrada.", "info")
    return redirect(url_for("login"))


# ---------------------------
# Routes - punch
# ---------------------------
@app.route("/ponto", methods=["GET", "POST"])
def punch():
    preview = None
    whatsapp_links = {}

    if request.method == "POST":
        registration = request.form.get("registration", "").strip()
        pin = request.form.get("pin", "").strip()

        employee = query_db("SELECT * FROM employees WHERE registration = ? AND active = 1", (registration,), one=True)
        if not employee:
            flash("Matrícula não encontrada.", "danger")
            return render_template("punch.html", preview=preview, current_time=now_dt(), whatsapp_links=whatsapp_links)

        if not check_password_hash(employee["pin_hash"], pin):
            flash("PIN inválido.", "danger")
            return render_template("punch.html", preview=preview, current_time=now_dt(), whatsapp_links=whatsapp_links)

        next_type = next_record_type(employee["id"])
        if next_type == "COMPLETO":
            flash("Todas as marcações do dia já foram registradas para este funcionário.", "warning")
            return render_template("punch.html", preview=preview, current_time=now_dt(), whatsapp_links=whatsapp_links)

        moment = now_iso()
        execute_db(
            """
            INSERT INTO time_records (employee_id, record_type, record_time, notes, ip_origin, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (employee["id"], next_type, moment, "Registro automático", request.remote_addr or "local", now_iso()),
        )
        preview = calculate_day_summary(employee["id"], today_str())
        log_action("REGISTRO_PONTO", f"{employee['name']} registrou {next_type}")
        flash(f"{next_type.replace('_', ' ').title()} registrada com sucesso para {employee['name']}.", "success")

        human_type = next_type.replace("_", " ").title()
        message = (
            f"Pet & Gatô - comprovante de ponto%0A"
            f"Funcionário: {employee['name']}%0A"
            f"Matrícula: {employee['registration']}%0A"
            f"Marcação: {human_type}%0A"
            f"Data/Hora: {datetime.strptime(moment, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M:%S')}"
        )
        employee_phone = employee["phone"] or ""
        manager_phone = get_setting("company_whatsapp", "")
        employee_link = build_whatsapp_link(employee_phone, message)
        manager_link = build_whatsapp_link(manager_phone, message)
        if employee_link:
            whatsapp_links["employee"] = employee_link
        if manager_link:
            whatsapp_links["manager"] = manager_link

    return render_template("punch.html", preview=preview, current_time=now_dt(), whatsapp_links=whatsapp_links)



# =========================================================
# Dashboard profissional: banco de horas, atrasos e ranking
# =========================================================

def safe_parse_dt(value):
    """Aceita datetime do PostgreSQL ou texto do SQLite."""
    if isinstance(value, datetime):
        return value
    return parse_dt(str(value))


def minutes_to_decimal_hours(minutes: int) -> float:
    return round((minutes or 0) / 60, 2)


def build_professional_dashboard(year: int, month: int) -> Dict[str, Any]:
    start, end = month_bounds(year, month)
    start_date = parse_date(start)
    end_date = parse_date(end)

    employees = query_db("SELECT * FROM employees WHERE active = 1 ORDER BY name")

    employee_rows = []
    ranking_extra = []
    ranking_delay = []
    daily_map: Dict[str, Dict[str, int]] = {}

    total_worked = 0
    total_scheduled = 0
    total_extra = 0
    total_negative = 0
    total_delay = 0
    total_early_exit = 0
    total_inconsistencies = 0
    total_records = 0

    current = start_date
    while current < end_date:
        key = current.strftime("%d/%m")
        daily_map[key] = {
            "extra": 0,
            "delay": 0,
            "negative": 0,
            "worked": 0,
        }
        current += timedelta(days=1)

    for emp in employees:
        emp_worked = 0
        emp_scheduled = 0
        emp_extra = 0
        emp_negative = 0
        emp_delay = 0
        emp_early_exit = 0
        emp_inconsistencies = 0
        emp_days_with_records = 0
        emp_complete_days = 0

        current = start_date
        while current < end_date:
            day_str = current.strftime("%Y-%m-%d")
            day_key = current.strftime("%d/%m")
            summary = calculate_day_summary(emp["id"], day_str)

            worked = int(summary.get("worked_minutes") or 0)
            scheduled = int(summary.get("scheduled_minutes") or 0)
            delay = int(summary.get("delay_minutes") or 0)
            early = int(summary.get("early_exit_minutes") or 0)
            extra_balance = int(summary.get("extra_minutes") or 0)
            records_count = len(summary.get("records") or [])

            if records_count > 0:
                emp_days_with_records += 1
                total_records += records_count

            if records_count == 4:
                emp_complete_days += 1

            if summary.get("inconsistency"):
                emp_inconsistencies += 1

            positive_extra = max(extra_balance, 0)
            negative_balance = abs(min(extra_balance, 0))

            emp_worked += worked
            emp_scheduled += scheduled
            emp_extra += positive_extra
            emp_negative += negative_balance
            emp_delay += delay
            emp_early_exit += early

            daily_map[day_key]["extra"] += positive_extra
            daily_map[day_key]["delay"] += delay
            daily_map[day_key]["negative"] += negative_balance
            daily_map[day_key]["worked"] += worked

            current += timedelta(days=1)

        total_worked += emp_worked
        total_scheduled += emp_scheduled
        total_extra += emp_extra
        total_negative += emp_negative
        total_delay += emp_delay
        total_early_exit += emp_early_exit
        total_inconsistencies += emp_inconsistencies

        bank_balance = emp_extra - emp_negative

        row = {
            "id": emp["id"],
            "name": emp["name"],
            "registration": emp["registration"],
            "worked_minutes": emp_worked,
            "scheduled_minutes": emp_scheduled,
            "extra_minutes": emp_extra,
            "negative_minutes": emp_negative,
            "bank_balance_minutes": bank_balance,
            "delay_minutes": emp_delay,
            "early_exit_minutes": emp_early_exit,
            "inconsistencies": emp_inconsistencies,
            "days_with_records": emp_days_with_records,
            "complete_days": emp_complete_days,
            "worked_h": format_minutes(emp_worked),
            "scheduled_h": format_minutes(emp_scheduled),
            "extra_h": format_minutes(emp_extra),
            "negative_h": format_minutes(emp_negative),
            "bank_balance_h": format_minutes(bank_balance),
            "delay_h": format_minutes(emp_delay),
            "early_exit_h": format_minutes(emp_early_exit),
        }
        employee_rows.append(row)
        ranking_extra.append(row)
        ranking_delay.append(row)

    ranking_extra = sorted(ranking_extra, key=lambda r: r["extra_minutes"], reverse=True)[:10]
    ranking_delay = sorted(ranking_delay, key=lambda r: r["delay_minutes"], reverse=True)[:10]
    employee_rows = sorted(employee_rows, key=lambda r: r["bank_balance_minutes"], reverse=True)

    chart_labels = list(daily_map.keys())
    chart_extra = [minutes_to_decimal_hours(daily_map[k]["extra"]) for k in chart_labels]
    chart_delay = [minutes_to_decimal_hours(daily_map[k]["delay"]) for k in chart_labels]
    chart_negative = [minutes_to_decimal_hours(daily_map[k]["negative"]) for k in chart_labels]
    chart_worked = [minutes_to_decimal_hours(daily_map[k]["worked"]) for k in chart_labels]

    return {
        "year": year,
        "month": month,
        "total_employees": len(employees),
        "total_records": total_records,
        "total_worked_minutes": total_worked,
        "total_scheduled_minutes": total_scheduled,
        "total_extra_minutes": total_extra,
        "total_negative_minutes": total_negative,
        "total_bank_balance_minutes": total_extra - total_negative,
        "total_delay_minutes": total_delay,
        "total_early_exit_minutes": total_early_exit,
        "total_inconsistencies": total_inconsistencies,
        "total_worked_h": format_minutes(total_worked),
        "total_scheduled_h": format_minutes(total_scheduled),
        "total_extra_h": format_minutes(total_extra),
        "total_negative_h": format_minutes(total_negative),
        "total_bank_balance_h": format_minutes(total_extra - total_negative),
        "total_delay_h": format_minutes(total_delay),
        "total_early_exit_h": format_minutes(total_early_exit),
        "employee_rows": employee_rows,
        "ranking_extra": ranking_extra,
        "ranking_delay": ranking_delay,
        "chart_labels": chart_labels,
        "chart_extra": chart_extra,
        "chart_delay": chart_delay,
        "chart_negative": chart_negative,
        "chart_worked": chart_worked,
    }


# ---------------------------
# Routes - dashboard/admin
# ---------------------------

@app.route("/dashboard")
@admin_required
def dashboard():
    today = today_str()
    today_date = date.today()
    year = request.args.get("year", type=int) or today_date.year
    month = request.args.get("month", type=int) or today_date.month

    total_employees = query_db("SELECT COUNT(*) AS total FROM employees WHERE active = 1", one=True)["total"]
    present_rows = query_db("SELECT COUNT(DISTINCT employee_id) AS total FROM time_records WHERE date(record_time) = ?", (today,), one=True)
    present_today = present_rows["total"]

    recent_records = query_db(
        """
        SELECT tr.*, e.name AS employee_name, e.registration
        FROM time_records tr
        JOIN employees e ON e.id = tr.employee_id
        ORDER BY tr.record_time DESC
        LIMIT 10
        """
    )

    employees = query_db("SELECT id, name FROM employees WHERE active = 1 ORDER BY name")

    delayed_count = 0
    inconsistent_count = 0
    for employee in employees:
        summary = calculate_day_summary(employee["id"], today)
        if summary["delay_minutes"] > 0:
            delayed_count += 1
        if summary["inconsistency"]:
            inconsistent_count += 1

    professional = build_professional_dashboard(year, month)

    return render_template(
        "dashboard.html",
        total_employees=total_employees,
        present_today=present_today,
        absent_today=max(total_employees - present_today, 0),
        delayed_count=delayed_count,
        inconsistent_count=inconsistent_count,
        recent_records=recent_records,
        format_dt=format_dt,
        professional=professional,
        selected_year=year,
        selected_month=month,
    )


@app.route("/funcionarios", methods=["GET", "POST"])
@admin_required
def employees():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        cpf = request.form.get("cpf", "").strip()
        registration = request.form.get("registration", "").strip()
        role_name = request.form.get("role_name", "").strip()
        phone = request.form.get("phone", "").strip()
        admission_date = request.form.get("admission_date", "").strip() or None
        pin = request.form.get("pin", "").strip()
        schedule_id = request.form.get("schedule_id", type=int)
        schedule_tue_fri = request.form.get("schedule_tue_fri", type=int)
        schedule_sat = request.form.get("schedule_sat", type=int)

        if not all([name, registration, pin]):
            flash("Nome, matrícula e PIN são obrigatórios.", "danger")
        else:
            try:
                employee_id = execute_db(
                    """
                    INSERT INTO employees (name, cpf, registration, role_name, phone, admission_date, pin_hash, schedule_id, active, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                    """,
                    (name, cpf, registration, role_name, phone, admission_date, generate_password_hash(pin), schedule_id, now_iso()),
                )
                for weekday in [1, 2, 3, 4]:
                    if schedule_tue_fri:
                        set_employee_week_schedule(employee_id, weekday, schedule_tue_fri)
                if schedule_sat:
                    set_employee_week_schedule(employee_id, 5, schedule_sat)
                log_action("CADASTRO_FUNCIONARIO", f"Funcionário {name} cadastrado")
                flash("Funcionário cadastrado com sucesso.", "success")
                return redirect(url_for("employees"))
            except sqlite3.IntegrityError:
                flash("A matrícula já existe. Use outra matrícula.", "danger")

    employee_list = query_db(
        """
        SELECT e.*, s.description AS schedule_description
        FROM employees e
        LEFT JOIN schedules s ON s.id = e.schedule_id
        ORDER BY e.name
        """
    )
    schedules = query_db("SELECT * FROM schedules WHERE active = 1 ORDER BY description")
    week_maps = {emp["id"]: get_employee_week_schedule_map(emp["id"]) for emp in employee_list}
    return render_template("employees.html", employees=employee_list, schedules=schedules, week_maps=week_maps, weekday_name=weekday_name)


@app.route("/funcionarios/<int:employee_id>/toggle")
@admin_required
def toggle_employee(employee_id: int):
    employee = query_db("SELECT * FROM employees WHERE id = ?", (employee_id,), one=True)
    if employee:
        new_active = 0 if employee["active"] else 1
        execute_db("UPDATE employees SET active = ? WHERE id = ?", (new_active, employee_id))
        log_action("ATUALIZACAO_FUNCIONARIO", f"Funcionário {employee['name']} ativo={new_active}")
        flash("Status do funcionário atualizado.", "success")
    return redirect(url_for("employees"))


@app.route("/funcionarios/<int:employee_id>/jornadas", methods=["POST"])
@admin_required
def update_employee_schedules(employee_id: int):
    employee = query_db("SELECT * FROM employees WHERE id = ?", (employee_id,), one=True)
    if not employee:
        flash("Funcionário não encontrado.", "danger")
        return redirect(url_for("employees"))

    default_schedule_id = request.form.get("schedule_id", type=int)
    tue_fri = request.form.get("schedule_tue_fri", type=int)
    saturday = request.form.get("schedule_sat", type=int)
    execute_db("UPDATE employees SET schedule_id = ? WHERE id = ?", (default_schedule_id, employee_id))
    for weekday in range(7):
        set_employee_week_schedule(employee_id, weekday, None)
    for weekday in [1, 2, 3, 4]:
        if tue_fri:
            set_employee_week_schedule(employee_id, weekday, tue_fri)
    if saturday:
        set_employee_week_schedule(employee_id, 5, saturday)

    log_action("ATUALIZACAO_JORNADA_FUNCIONARIO", f"Jornadas atualizadas para {employee['name']}")
    flash("Jornadas do funcionário atualizadas.", "success")
    return redirect(url_for("employees"))


# =========================================================
# Registros profissionais: visão por dia + ajuste seguro
# =========================================================

def _time_value_from_record(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        return parse_dt(str(value)).strftime("%H:%M")
    except Exception:
        text_value = str(value)
        if len(text_value) >= 16:
            return text_value[11:16]
        return ""


def _parse_time_input(ref_date: str, time_value: str) -> Optional[str]:
    time_value = (time_value or "").strip()
    if not time_value:
        return None
    try:
        if len(time_value) == 5:
            datetime.strptime(time_value, "%H:%M")
            return f"{ref_date} {time_value}:00"
        if len(time_value) == 8:
            datetime.strptime(time_value, "%H:%M:%S")
            return f"{ref_date} {time_value}"
    except ValueError:
        return None
    return None


def _validate_day_times(values: Dict[str, Optional[str]]) -> Optional[str]:
    parsed = {key: parse_dt(value) for key, value in values.items() if value}
    entrada = parsed.get("ENTRADA")
    saida_intervalo = parsed.get("SAIDA_INTERVALO")
    retorno_intervalo = parsed.get("RETORNO_INTERVALO")
    saida = parsed.get("SAIDA")

    if saida and entrada and saida <= entrada:
        return "A saída final precisa ser maior que a entrada."
    if entrada and saida_intervalo and saida_intervalo <= entrada:
        return "A saída para intervalo precisa ser maior que a entrada."
    if saida_intervalo and retorno_intervalo and retorno_intervalo <= saida_intervalo:
        return "O retorno do intervalo precisa ser maior que a saída para intervalo."
    if retorno_intervalo and saida and saida <= retorno_intervalo:
        return "A saída final precisa ser maior que o retorno do intervalo."
    return None


def _get_record_by_type(employee_id: int, ref_date: str, record_type: str):
    return query_db(
        """
        SELECT * FROM time_records
        WHERE employee_id = ? AND date(record_time) = ? AND record_type = ?
        ORDER BY id ASC
        LIMIT 1
        """,
        (employee_id, ref_date, record_type),
        one=True,
    )


def _build_records_group(employee: Any, ref_date: str) -> Dict[str, Any]:
    summary = calculate_day_summary(employee["id"], ref_date)
    sequence = summary.get("sequence") or {}
    times = {
        "ENTRADA": _time_value_from_record(sequence.get("ENTRADA")),
        "SAIDA_INTERVALO": _time_value_from_record(sequence.get("SAIDA_INTERVALO")),
        "RETORNO_INTERVALO": _time_value_from_record(sequence.get("RETORNO_INTERVALO")),
        "SAIDA": _time_value_from_record(sequence.get("SAIDA")),
    }
    return {
        "employee": employee,
        "schedule": summary.get("schedule"),
        "weekday_name": summary.get("weekday_name"),
        "times": times,
        "worked_h": format_minutes(int(summary.get("worked_minutes") or 0)),
        "delay_h": format_minutes(int(summary.get("delay_minutes") or 0)),
        "extra_h": format_minutes(int(summary.get("extra_minutes") or 0)),
        "delay_minutes": int(summary.get("delay_minutes") or 0),
        "extra_minutes": int(summary.get("extra_minutes") or 0),
        "inconsistency": bool(summary.get("inconsistency")),
    }


@app.route("/registros")
@admin_required
def records():
    employee_id = request.args.get("employee_id", type=int)
    ref_date = request.args.get("ref_date", today_str())

    try:
        parse_date(ref_date)
    except Exception:
        ref_date = today_str()

    employees = query_db("SELECT id, name, registration FROM employees WHERE active = 1 ORDER BY name")
    selected_employees = employees
    if employee_id:
        selected_employees = [emp for emp in employees if int(emp["id"]) == int(employee_id)]

    grouped_days = [_build_records_group(emp, ref_date) for emp in selected_employees]

    params = [ref_date]
    where = ["date(tr.record_time) = ?"]
    if employee_id:
        where.append("tr.employee_id = ?")
        params.append(employee_id)

    record_list = query_db(
        f"""
        SELECT tr.*, e.name AS employee_name, e.registration
        FROM time_records tr
        JOIN employees e ON e.id = tr.employee_id
        WHERE {' AND '.join(where)}
        ORDER BY e.name ASC, tr.record_time ASC, tr.id ASC
        """,
        tuple(params),
    )

    return render_template(
        "records.html",
        records=record_list,
        employees=employees,
        grouped_days=grouped_days,
        ref_date=ref_date,
        ref_date_br=parse_date(ref_date).strftime("%d/%m/%Y"),
        employee_id=employee_id,
        format_dt=format_dt,
    )


@app.route("/registros/dia/<int:employee_id>/<ref_date>/ajustar", methods=["POST"])
@admin_required
def update_day_records(employee_id: int, ref_date: str):
    employee = query_db("SELECT * FROM employees WHERE id = ?", (employee_id,), one=True)
    if not employee:
        flash("Funcionário não encontrado.", "danger")
        return redirect(url_for("records", ref_date=ref_date))

    try:
        parse_date(ref_date)
    except Exception:
        flash("Data inválida para ajuste.", "danger")
        return redirect(url_for("records"))

    reason = request.form.get("reason", "").strip()
    if not reason:
        flash("Informe o motivo do ajuste.", "danger")
        return redirect(url_for("records", ref_date=ref_date, employee_id=employee_id))

    values: Dict[str, Optional[str]] = {}
    for record_type in record_sequence():
        raw_time = request.form.get(record_type, "").strip()
        if raw_time:
            parsed_time = _parse_time_input(ref_date, raw_time)
            if not parsed_time:
                flash(f"Horário inválido em {record_type}. Use HH:MM.", "danger")
                return redirect(url_for("records", ref_date=ref_date, employee_id=employee_id))
            values[record_type] = parsed_time
        else:
            values[record_type] = None

    validation_error = _validate_day_times(values)
    if validation_error:
        flash(validation_error, "danger")
        return redirect(url_for("records", ref_date=ref_date, employee_id=employee_id))

    changed = 0
    for record_type in record_sequence():
        new_value = values.get(record_type)
        if not new_value:
            continue

        existing = _get_record_by_type(employee_id, ref_date, record_type)
        if existing:
            old_value = existing["record_time"]
            if str(old_value) == str(new_value):
                continue
            execute_db("UPDATE time_records SET record_time = ?, notes = ? WHERE id = ?", (new_value, "Ajuste manual", existing["id"]))
            execute_db(
                """
                INSERT INTO time_adjustments (record_id, employee_id, admin_user_id, old_value, new_value, reason, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (existing["id"], employee_id, session["user_id"], str(old_value), new_value, reason, now_iso()),
            )
            changed += 1
        else:
            new_id = execute_db(
                """
                INSERT INTO time_records (employee_id, record_type, record_time, notes, ip_origin, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (employee_id, record_type, new_value, "Criado por ajuste manual", get_client_ip(), now_iso()),
            )
            execute_db(
                """
                INSERT INTO time_adjustments (record_id, employee_id, admin_user_id, old_value, new_value, reason, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (new_id, employee_id, session["user_id"], "", new_value, reason, now_iso()),
            )
            changed += 1

    log_action("AJUSTE_JORNADA_DIA", f"{employee['name']} em {ref_date}. Alterações: {changed}. Motivo: {reason}")
    flash(f"Jornada ajustada com sucesso. Alterações aplicadas: {changed}.", "success")
    return redirect(url_for("records", ref_date=ref_date, employee_id=employee_id))

# =========================================================
# Fim registros profissionais
# =========================================================
@app.route("/registros/<int:record_id>/ajustar", methods=["POST"])
@admin_required
def adjust_record(record_id: int):
    record = query_db("SELECT * FROM time_records WHERE id = ?", (record_id,), one=True)
    if not record:
        flash("Registro não encontrado.", "danger")
        return redirect(url_for("records"))

    new_value = request.form.get("new_value", "").strip()
    reason = request.form.get("reason", "").strip()
    if not new_value or not reason:
        flash("Nova data/hora e motivo são obrigatórios.", "danger")
        return redirect(url_for("records", ref_date=parse_dt(record["record_time"]).strftime("%Y-%m-%d")))

    try:
        datetime.strptime(new_value, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        flash("Formato inválido. Use YYYY-MM-DD HH:MM:SS.", "danger")
        return redirect(url_for("records", ref_date=parse_dt(record["record_time"]).strftime("%Y-%m-%d")))

    execute_db(
        "INSERT INTO time_adjustments (record_id, employee_id, admin_user_id, old_value, new_value, reason, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (record_id, record["employee_id"], session["user_id"], record["record_time"], new_value, reason, now_iso()),
    )
    execute_db("UPDATE time_records SET record_time = ? WHERE id = ?", (new_value, record_id))
    log_action("AJUSTE_REGISTRO", f"Registro {record_id} ajustado. Motivo: {reason}")
    flash("Registro ajustado com sucesso.", "success")
    return redirect(url_for("records", ref_date=parse_dt(new_value).strftime("%Y-%m-%d"), employee_id=record["employee_id"]))


@app.route("/relatorios", methods=["GET", "POST"])
@admin_required
def reports():
    employees = query_db("SELECT id, name, registration FROM employees WHERE active = 1 ORDER BY name")
    today = date.today()

    if request.method == "POST":
        action = request.form.get("action")
        year = request.form.get("year", type=int) or today.year
        month = request.form.get("month", type=int) or today.month
        employee_id = request.form.get("employee_id", type=int)

        if action == "excel":
            file_path = generate_excel_report(year, month)
            log_action("EXPORTAR_EXCEL", f"Relatório mensal {month:02d}/{year}")
            return send_file(file_path, as_attachment=True)

        if action == "pdf":
            if not employee_id:
                flash("Selecione um funcionário para gerar o PDF.", "danger")
            else:
                file_path = generate_pdf_report(employee_id, year, month)
                log_action("EXPORTAR_PDF", f"Espelho de ponto do funcionário {employee_id} - {month:02d}/{year}")
                return send_file(file_path, as_attachment=True)

        if action == "backup":
            backup_path = create_backup()
            log_action("BACKUP", f"Backup gerado: {backup_path.name}")
            return send_file(backup_path, as_attachment=True)

    return render_template("reports.html", employees=employees, today=today)


@app.route("/jornadas", methods=["GET", "POST"])
@admin_required
def schedules():
    if request.method == "POST":
        description = request.form.get("description", "").strip()
        entry_time = request.form.get("entry_time", "").strip()
        lunch_out_time = request.form.get("lunch_out_time", "").strip()
        lunch_in_time = request.form.get("lunch_in_time", "").strip()
        exit_time = request.form.get("exit_time", "").strip()
        tolerance_minutes = request.form.get("tolerance_minutes", type=int) or 10

        if all([description, entry_time, lunch_out_time, lunch_in_time, exit_time]):
            execute_db(
                """
                INSERT INTO schedules (description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes, active)
                VALUES (?, ?, ?, ?, ?, ?, 1)
                """,
                (description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes),
            )
            log_action("CADASTRO_JORNADA", f"Jornada cadastrada: {description}")
            flash("Jornada cadastrada com sucesso.", "success")
            return redirect(url_for("schedules"))
        flash("Preencha todos os campos da jornada.", "danger")

    items = query_db("SELECT * FROM schedules ORDER BY id DESC")
    return render_template("schedules.html", schedules=items)


@app.route("/configuracoes", methods=["GET", "POST"])
@admin_required
def settings_page():
    if request.method == "POST":
        company_whatsapp = request.form.get("company_whatsapp", "").strip()
        set_setting("company_whatsapp", company_whatsapp)
        log_action("CONFIGURACOES", "WhatsApp da empresa atualizado")
        flash("Configurações salvas com sucesso.", "success")
        return redirect(url_for("settings_page"))

    return render_template("settings.html", company_whatsapp=get_setting("company_whatsapp", ""))


@app.route("/health")
def health():
    return {"status": "ok", "time": now_iso()}


@app.route("/sw.js")
def service_worker():
    response = send_from_directory(BASE_DIR / "static", "sw.js")
    response.headers["Cache-Control"] = "no-cache"
    response.headers["Service-Worker-Allowed"] = "/"
    response.mimetype = "application/javascript"
    return response


@app.context_processor
def inject_globals():
    return {
        "session": session,
        "now": now_dt,
        "format_dt": format_dt,
    }


# ===========================
# Notificações profissionais
# ===========================
_NOTIFICATIONS_DB_READY = False


def _table_exists(table_name: str) -> bool:
    if os.getenv("DATABASE_URL"):
        row = query_db(
            "SELECT to_regclass(?) AS table_name",
            (table_name,),
            one=True,
        )
        return bool(row and row["table_name"])
    row = query_db(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
        (table_name,),
        one=True,
    )
    return bool(row)


def ensure_notifications_tables() -> None:
    """Cria tabelas de notificações/feriados sem alterar dados existentes."""
    if os.getenv("DATABASE_URL"):
        execute_db("""
            CREATE TABLE IF NOT EXISTS holidays (
                id SERIAL PRIMARY KEY,
                holiday_date TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                kind TEXT NOT NULL DEFAULT 'Nacional',
                created_at TEXT NOT NULL
            )
        """)
        execute_db("""
            CREATE TABLE IF NOT EXISTS notification_logs (
                id SERIAL PRIMARY KEY,
                unique_key TEXT NOT NULL UNIQUE,
                alert_type TEXT NOT NULL,
                employee_id INTEGER,
                employee_name TEXT,
                message TEXT NOT NULL,
                created_at TEXT NOT NULL,
                sent_status TEXT NOT NULL DEFAULT 'PENDENTE'
            )
        """)
    else:
        execute_db("""
            CREATE TABLE IF NOT EXISTS holidays (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                holiday_date TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                kind TEXT NOT NULL DEFAULT 'Nacional',
                created_at TEXT NOT NULL
            )
        """)
        execute_db("""
            CREATE TABLE IF NOT EXISTS notification_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                unique_key TEXT NOT NULL UNIQUE,
                alert_type TEXT NOT NULL,
                employee_id INTEGER,
                employee_name TEXT,
                message TEXT NOT NULL,
                created_at TEXT NOT NULL,
                sent_status TEXT NOT NULL DEFAULT 'PENDENTE'
            )
        """)

    # Configurações padrão
    for key in ("notify_phone_1", "notify_phone_2"):
        row = query_db("SELECT value FROM settings WHERE key = ?", (key,), one=True)
        if not row:
            set_setting(key, "")

    seed_holidays_2026()


def ensure_notifications_ready() -> None:
    global _NOTIFICATIONS_DB_READY
    if not _NOTIFICATIONS_DB_READY:
        ensure_notifications_tables()
        _NOTIFICATIONS_DB_READY = True


@app.before_request
def notifications_bootstrap():
    # Garante as tabelas sem bloquear static/health.
    if request.endpoint not in ("static", "health", "service_worker"):
        try:
            ensure_notifications_ready()
        except Exception as exc:
            print("Erro ao inicializar notificações:", exc)


def seed_holidays_2026() -> None:
    """Feriados/pontos facultativos iniciais de 2026. Pode editar pelo painel."""
    holidays = [
        ("2026-01-01", "Confraternização Universal", "Nacional"),
        ("2026-02-16", "Carnaval", "Ponto facultativo"),
        ("2026-02-17", "Carnaval", "Ponto facultativo"),
        ("2026-02-18", "Quarta-feira de Cinzas", "Ponto facultativo"),
        ("2026-04-03", "Paixão de Cristo / Sexta-feira Santa", "Nacional/Municipal"),
        ("2026-04-21", "Tiradentes", "Nacional"),
        ("2026-05-01", "Dia do Trabalhador", "Nacional"),
        ("2026-06-04", "Corpus Christi", "Ponto facultativo/Municipal"),
        ("2026-09-07", "Independência do Brasil", "Nacional"),
        ("2026-10-12", "Nossa Senhora Aparecida", "Nacional"),
        ("2026-11-02", "Finados", "Nacional"),
        ("2026-11-15", "Proclamação da República", "Nacional"),
        ("2026-11-20", "Consciência Negra", "Nacional"),
        ("2026-12-25", "Natal", "Nacional"),
    ]
    for holiday_date, name, kind in holidays:
        try:
            execute_db(
                """
                INSERT INTO holidays (holiday_date, name, kind, created_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(holiday_date) DO NOTHING
                """,
                (holiday_date, name, kind, now_iso()),
            )
        except Exception:
            pass


def get_holiday(ref_date: str):
    ensure_notifications_ready()
    return query_db(
        "SELECT * FROM holidays WHERE holiday_date = ?",
        (ref_date,),
        one=True,
    )


def is_holiday_date(ref_date: str) -> bool:
    return bool(get_holiday(ref_date))


def create_notification(unique_key: str, alert_type: str, message: str, employee_id=None, employee_name=None) -> None:
    ensure_notifications_ready()
    try:
        execute_db(
            """
            INSERT INTO notification_logs (unique_key, alert_type, employee_id, employee_name, message, created_at, sent_status)
            VALUES (?, ?, ?, ?, ?, ?, 'PENDENTE')
            ON CONFLICT(unique_key) DO NOTHING
            """,
            (unique_key, alert_type, employee_id, employee_name, message, now_iso()),
        )
    except Exception as exc:
        print("Erro ao criar notificação:", exc)


def wa_link(phone: str, message: str) -> Optional[str]:
    phone = normalize_phone(phone or "")
    if not phone:
        return None
    return build_whatsapp_link(phone, message)


def check_attendance_notifications(ref_date: Optional[str] = None) -> int:
    """Gera alertas de atraso e batidas esquecidas."""
    ensure_notifications_ready()
    ref_date = ref_date or today_str()
    holiday = get_holiday(ref_date)
    if holiday:
        create_notification(
            f"{ref_date}:FERIADO",
            "FERIADO",
            f"Hoje é feriado: {holiday['name']} ({holiday['kind']}). O sistema não gerou cobrança automática de atraso.",
        )
        return 1

    generated = 0
    now_value = now_dt()
    employees_rows = query_db("SELECT * FROM employees WHERE active = 1 ORDER BY name")

    for employee in employees_rows:
        schedule = get_schedule_for_day(employee["id"], ref_date)
        if not schedule:
            continue

        summary = calculate_day_summary(employee["id"], ref_date)
        seq = summary["sequence"]
        tolerance = int(schedule["tolerance_minutes"] or 0)
        ref_day = parse_date(ref_date)

        checks = []

        # Alguns funcionários antigos podem estar sem jornada completa no banco.
        # Nesses casos, ignoramos o horário faltante para evitar erro 500.
        if schedule.get("entry_time"):
            checks.append((
                "ATRASO_ENTRADA",
                "ENTRADA",
                combine_day_time(ref_day, schedule["entry_time"]) + timedelta(minutes=tolerance),
                f"{employee['name']} ainda não registrou ENTRADA. Jornada prevista: {schedule['entry_time']}.",
            ))

        if schedule.get("lunch_out_time"):
            checks.append((
                "ESQUECEU_SAIDA_INTERVALO",
                "SAIDA_INTERVALO",
                combine_day_time(ref_day, schedule["lunch_out_time"]) + timedelta(minutes=tolerance),
                f"{employee['name']} registrou entrada, mas ainda não registrou SAÍDA PARA INTERVALO. Horário previsto: {schedule['lunch_out_time']}.",
            ))

        if schedule.get("lunch_in_time"):
            checks.append((
                "ESQUECEU_RETORNO_INTERVALO",
                "RETORNO_INTERVALO",
                combine_day_time(ref_day, schedule["lunch_in_time"]) + timedelta(minutes=tolerance),
                f"{employee['name']} saiu para intervalo, mas ainda não registrou RETORNO DO INTERVALO. Horário previsto: {schedule['lunch_in_time']}.",
            ))

        if schedule.get("exit_time"):
            checks.append((
                "ESQUECEU_SAIDA",
                "SAIDA",
                combine_day_time(ref_day, schedule["exit_time"]) + timedelta(minutes=tolerance),
                f"{employee['name']} ainda não registrou SAÍDA. Horário previsto: {schedule['exit_time']}.",
            ))


        for alert_type, required_record, limit_time, message in checks:
            if now_value < limit_time:
                continue

            if required_record == "SAIDA_INTERVALO" and not seq.get("ENTRADA"):
                continue
            if required_record == "RETORNO_INTERVALO" and not seq.get("SAIDA_INTERVALO"):
                continue
            if required_record == "SAIDA" and not seq.get("ENTRADA"):
                continue

            if not seq.get(required_record):
                unique = f"{ref_date}:{employee['id']}:{alert_type}"
                full_message = f"Alerta Pet & Gatô%0AData: {parse_date(ref_date).strftime('%d/%m/%Y')}%0A{message}"
                create_notification(unique, alert_type, full_message, employee["id"], employee["name"])
                generated += 1

    return generated


def create_month_closing_notification(year: Optional[int] = None, month: Optional[int] = None) -> None:
    ensure_notifications_ready()
    today = date.today()
    year = year or today.year
    month = month or today.month
    start, end = month_bounds(year, month)

    employees_rows = query_db("SELECT id, name FROM employees WHERE active = 1 ORDER BY name")
    total_extra = 0
    total_delay = 0
    total_inconsistencies = 0

    for employee in employees_rows:
        current = parse_date(start)
        end_date = parse_date(end)
        while current < end_date:
            summary = calculate_day_summary(employee["id"], current.strftime("%Y-%m-%d"))
            if summary["extra_minutes"] > 0:
                total_extra += summary["extra_minutes"]
            total_delay += summary["delay_minutes"]
            if summary["inconsistency"]:
                total_inconsistencies += 1
            current += timedelta(days=1)

    msg = (
        f"Fechamento mensal Pet & Gatô%0A"
        f"Referência: {month:02d}/{year}%0A"
        f"Horas extras totais: {format_minutes(total_extra)}%0A"
        f"Atrasos totais: {format_minutes(total_delay)}%0A"
        f"Inconsistências: {total_inconsistencies}"
    )
    create_notification(f"{year}-{month:02d}:FECHAMENTO", "FECHAMENTO_MENSAL", msg)


@app.route("/notificacoes")
@admin_required
def notificacoes():
    ensure_notifications_ready()
    phone1 = get_setting("notify_phone_1", "")
    phone2 = get_setting("notify_phone_2", "")
    logs = query_db("SELECT * FROM notification_logs ORDER BY created_at DESC, id DESC LIMIT 100")
    prepared_logs = []
    for log in logs:
        row = dict(log)
        row["link1"] = wa_link(phone1, row["message"])
        row["link2"] = wa_link(phone2, row["message"])
        prepared_logs.append(row)

    holidays = query_db("SELECT * FROM holidays ORDER BY holiday_date ASC")
    return render_template(
        "notificacoes.html",
        notify_phone_1=phone1,
        notify_phone_2=phone2,
        logs=prepared_logs,
        holidays=holidays,
        holiday_today=get_holiday(today_str()),
    )



@app.route("/notificacoes/config", methods=["POST"])
@admin_required
def notifications_settings():
    ensure_notifications_ready()
    set_setting("notify_phone_1", request.form.get("notify_phone_1", "").strip())
    set_setting("notify_phone_2", request.form.get("notify_phone_2", "").strip())
    flash("Números de WhatsApp salvos com sucesso.", "success")
    return redirect(url_for("notificacoes"))


@app.route("/notificacoes/verificar", methods=["POST"])
@admin_required
def notifications_run_check():
    generated = check_attendance_notifications()
    flash(f"Verificação concluída. Alertas novos gerados: {generated}.", "success")
    return redirect(url_for("notificacoes"))


@app.route("/notificacoes/fechamento", methods=["POST"])
@admin_required
def notifications_month_close():
    create_month_closing_notification()
    flash("Fechamento mensal gerado em notificações.", "success")
    return redirect(url_for("notificacoes"))


@app.route("/notificacoes/feriados", methods=["POST"])
@admin_required
def notifications_add_holiday():
    ensure_notifications_ready()
    holiday_date = request.form.get("holiday_date", "").strip()
    holiday_name = request.form.get("holiday_name", "").strip()
    holiday_type = request.form.get("holiday_type", "Nacional").strip() or "Nacional"
    if not holiday_date or not holiday_name:
        flash("Informe data e nome do feriado.", "danger")
        return redirect(url_for("notificacoes"))

    execute_db(
        """
        INSERT INTO holidays (holiday_date, name, kind, created_at)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(holiday_date) DO UPDATE SET name = excluded.name, kind = excluded.kind
        """,
        (holiday_date, holiday_name, holiday_type, now_iso()),
    )
    flash("Feriado salvo com sucesso.", "success")
    return redirect(url_for("notificacoes"))



# =========================================================
# Multiempresa / SaaS - Pet & Gatô Ponto
# =========================================================
_MULTIEMPRESA_DB_READY = False

SYSTEM_COMMERCIAL_NAME = "PontoFácil Pro"
SYSTEM_COMMERCIAL_SUBTITLE = "Controle de ponto online com banco de horas, alertas e relatórios."

PLAN_PRICES = {
    "BASICO": 29.00,
    "PRO": 59.00,
    "PREMIUM": 99.00,
}

PLAN_LIMITS = {
    "BASICO": {"name": "Básico", "price": "R$29/mês", "employees": 5},
    "PRO": {"name": "Profissional", "price": "R$59/mês", "employees": 20},
    "PREMIUM": {"name": "Premium", "price": "R$99/mês", "employees": 9999},
}


def _column_exists(table_name: str, column_name: str) -> bool:
    if os.getenv("DATABASE_URL"):
        row = query_db(
            """
            SELECT column_name FROM information_schema.columns
            WHERE table_name = ? AND column_name = ?
            """,
            (table_name, column_name),
            one=True,
        )
        return bool(row)
    row = query_db(f"PRAGMA table_info({table_name})")
    return any(r["name"] == column_name for r in row)


def _add_company_column(table_name: str) -> None:
    if _column_exists(table_name, "company_id"):
        return
    execute_db(f"ALTER TABLE {table_name} ADD COLUMN company_id INTEGER")


def ensure_multiempresa_tables() -> None:
    """Prepara o sistema para vender para várias empresas sem apagar dados antigos."""
    if os.getenv("DATABASE_URL"):
        execute_db("""
            CREATE TABLE IF NOT EXISTS companies (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                slug TEXT NOT NULL UNIQUE,
                responsible_name TEXT,
                email TEXT,
                phone TEXT,
                plan TEXT NOT NULL DEFAULT 'BASICO',
                status TEXT NOT NULL DEFAULT 'ATIVO',
                trial_until TEXT,
                created_at TEXT NOT NULL
            )
        """)
        execute_db("""
            CREATE TABLE IF NOT EXISTS billing_events (
                id SERIAL PRIMARY KEY,
                company_id INTEGER,
                event_type TEXT NOT NULL,
                description TEXT,
                amount TEXT,
                status TEXT NOT NULL DEFAULT 'PENDENTE',
                created_at TEXT NOT NULL
            )
        """)
    else:
        execute_db("""
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                slug TEXT NOT NULL UNIQUE,
                responsible_name TEXT,
                email TEXT,
                phone TEXT,
                plan TEXT NOT NULL DEFAULT 'BASICO',
                status TEXT NOT NULL DEFAULT 'ATIVO',
                trial_until TEXT,
                created_at TEXT NOT NULL
            )
        """)
        execute_db("""
            CREATE TABLE IF NOT EXISTS billing_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER,
                event_type TEXT NOT NULL,
                description TEXT,
                amount TEXT,
                status TEXT NOT NULL DEFAULT 'PENDENTE',
                created_at TEXT NOT NULL
            )
        """)

    for table in [
        "users", "employees", "schedules", "employee_schedule_days", "time_records",
        "time_adjustments", "system_logs", "holidays", "notification_logs"
    ]:
        try:
            _add_company_column(table)
        except Exception as exc:
            print(f"Aviso multiempresa: não foi possível ajustar {table}: {exc}")

    default_company = query_db("SELECT * FROM companies WHERE slug = ?", ("pet-gato",), one=True)
    if not default_company:
        execute_db(
            """
            INSERT INTO companies (name, slug, responsible_name, email, phone, plan, status, trial_until, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("Pet & Gatô", "pet-gato", "Diego Poppe", "", "", "PREMIUM", "ATIVO", "", now_iso()),
        )

    default_company = query_db("SELECT * FROM companies WHERE slug = ?", ("pet-gato",), one=True)
    default_id = default_company["id"]

    for table in ["users", "employees", "schedules", "employee_schedule_days", "time_records", "time_adjustments", "system_logs"]:
        try:
            execute_db(f"UPDATE {table} SET company_id = ? WHERE company_id IS NULL", (default_id,))
        except Exception as exc:
            print(f"Aviso multiempresa update {table}: {exc}")


def ensure_multiempresa_ready() -> None:
    global _MULTIEMPRESA_DB_READY
    if not _MULTIEMPRESA_DB_READY:
        ensure_multiempresa_tables()
        _MULTIEMPRESA_DB_READY = True


@app.before_request
def multiempresa_bootstrap():
    if request.endpoint not in ("static", "health", "service_worker"):
        try:
            ensure_multiempresa_ready()
        except Exception as exc:
            print("Erro ao inicializar multiempresa:", exc)


def current_company_id() -> Optional[int]:
    cid = session.get("company_id")
    try:
        return int(cid) if cid else None
    except Exception:
        return None


def current_company():
    cid = current_company_id()
    if not cid:
        return None
    return query_db("SELECT * FROM companies WHERE id = ?", (cid,), one=True)


def is_super_admin() -> bool:
    return session.get("role") in ("super_admin", "master") or session.get("user_name") == "Administrador"


def company_plan_limit(company_id: Optional[int] = None) -> int:
    company = query_db("SELECT * FROM companies WHERE id = ?", (company_id or current_company_id(),), one=True)
    if not company:
        return 0
    plan = (company.get("plan") or "BASICO").upper()
    return PLAN_LIMITS.get(plan, PLAN_LIMITS["BASICO"])["employees"]


def active_employee_count(company_id: Optional[int] = None) -> int:
    row = query_db(
        "SELECT COUNT(*) AS total FROM employees WHERE active = 1 AND company_id = ?",
        (company_id or current_company_id(),),
        one=True,
    )
    return int(row["total"] or 0)


def check_employee_limit() -> bool:
    cid = current_company_id()
    if not cid:
        return True
    return active_employee_count(cid) < company_plan_limit(cid)


@app.context_processor
def inject_company_globals():
    return {
        "current_company": current_company,
        "PLAN_LIMITS": PLAN_LIMITS,
        "is_super_admin": is_super_admin,
    }


@app.route("/empresas", methods=["GET", "POST"])
@admin_required
def companies_page():
    ensure_multiempresa_ready()
    if not is_super_admin():
        flash("Acesso restrito ao administrador principal.", "danger")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        slug = request.form.get("slug", "").strip().lower().replace(" ", "-")
        responsible_name = request.form.get("responsible_name", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        plan = request.form.get("plan", "BASICO").strip().upper()
        status = request.form.get("status", "ATIVO").strip().upper()
        trial_until = request.form.get("trial_until", "").strip()

        if not name or not slug:
            flash("Nome e identificador da empresa são obrigatórios.", "danger")
        else:
            execute_db(
                """
                INSERT INTO companies (name, slug, responsible_name, email, phone, plan, status, trial_until, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(slug) DO UPDATE SET
                    name = excluded.name,
                    responsible_name = excluded.responsible_name,
                    email = excluded.email,
                    phone = excluded.phone,
                    plan = excluded.plan,
                    status = excluded.status,
                    trial_until = excluded.trial_until
                """,
                (name, slug, responsible_name, email, phone, plan, status, trial_until, now_iso()),
            )
            flash("Empresa salva com sucesso.", "success")
            return redirect(url_for("companies_page"))

    companies = query_db("SELECT * FROM companies ORDER BY created_at DESC, id DESC")
    rows = []
    for c in companies:
        row = dict(c)
        row["employees_count"] = active_employee_count(row["id"])
        row["limit"] = company_plan_limit(row["id"])
        rows.append(row)
    return render_template("companies.html", companies=rows)


@app.route("/empresas/<int:company_id>/entrar")
@admin_required
def switch_company(company_id: int):
    if not is_super_admin():
        flash("Acesso restrito.", "danger")
        return redirect(url_for("dashboard"))
    company = query_db("SELECT * FROM companies WHERE id = ?", (company_id,), one=True)
    if company:
        session["company_id"] = company["id"]
        session["company_name"] = company["name"]
        flash(f"Empresa ativa: {company['name']}", "success")
    return redirect(url_for("dashboard"))


@app.route("/planos")
@admin_required
def plans_page():
    ensure_multiempresa_ready()
    return render_template("plans.html", company=current_company())


@app.route("/assinatura")
@admin_required
def subscription_page():
    ensure_multiempresa_ready()
    company = current_company()
    events = []
    if company:
        events = query_db("SELECT * FROM billing_events WHERE company_id = ? ORDER BY created_at DESC", (company["id"],))
    return render_template("subscription.html", company=company, events=events)



# =========================================================
# Cobrança automática Mercado Pago - PontoFácil Pro
# =========================================================
_BILLING_DB_READY = False


def mp_access_token() -> str:
    return os.getenv("MP_ACCESS_TOKEN", "").strip()


def app_public_url() -> str:
    env_url = os.getenv("APP_PUBLIC_URL", "").strip().rstrip("/")
    if env_url:
        return env_url
    try:
        return request.url_root.rstrip("/")
    except Exception:
        return ""


def plan_price(plan_key: str) -> float:
    plan_key = (plan_key or "BASICO").upper()
    return float(PLAN_PRICES.get(plan_key, PLAN_PRICES["BASICO"]))


def plan_label(plan_key: str) -> str:
    plan_key = (plan_key or "BASICO").upper()
    data = PLAN_LIMITS.get(plan_key, PLAN_LIMITS["BASICO"])
    return data["name"]


def ensure_billing_tables() -> None:
    """Adiciona campos financeiros sem apagar dados existentes."""
    ensure_multiempresa_ready()

    for column in ["paid_until", "last_payment_at", "payment_status"]:
        try:
            if not _column_exists("companies", column):
                execute_db(f"ALTER TABLE companies ADD COLUMN {column} TEXT")
        except Exception as exc:
            print(f"Aviso cobrança: não foi possível criar companies.{column}: {exc}")

    for column in ["external_reference", "mp_preference_id", "mp_payment_id", "checkout_url", "plan", "paid_at"]:
        try:
            if not _column_exists("billing_events", column):
                execute_db(f"ALTER TABLE billing_events ADD COLUMN {column} TEXT")
        except Exception as exc:
            print(f"Aviso cobrança: não foi possível criar billing_events.{column}: {exc}")


def ensure_billing_ready() -> None:
    global _BILLING_DB_READY
    if not _BILLING_DB_READY:
        ensure_billing_tables()
        _BILLING_DB_READY = True


@app.before_request
def billing_bootstrap():
    if request.endpoint not in ("static", "health", "service_worker"):
        try:
            ensure_billing_ready()
        except Exception as exc:
            print("Erro ao inicializar cobrança:", exc)


def create_billing_event(company_id: int, plan_key: str) -> int:
    amount = f"{plan_price(plan_key):.2f}"
    description = f"Assinatura {SYSTEM_COMMERCIAL_NAME} - Plano {plan_label(plan_key)}"
    event_id = execute_db(
        """
        INSERT INTO billing_events (company_id, event_type, description, amount, status, created_at, plan)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (company_id, "ASSINATURA", description, amount, "AGUARDANDO_PAGAMENTO", now_iso(), plan_key.upper()),
    )
    external_reference = f"company:{company_id}:plan:{plan_key.upper()}:event:{event_id}"
    execute_db("UPDATE billing_events SET external_reference = ? WHERE id = ?", (external_reference, event_id))
    return int(event_id)


def create_mercadopago_preference(company: Any, plan_key: str, event_id: int) -> Dict[str, Any]:
    """Cria uma preferência do Checkout Pro do Mercado Pago.

    Fluxo recomendado para SaaS:
    - o cliente clica em Pagar agora;
    - o Mercado Pago abre uma página segura com Pix e cartão;
    - o webhook confirma o pagamento e libera a empresa automaticamente.
    """
    token = mp_access_token()
    if not token:
        raise RuntimeError("MP_ACCESS_TOKEN não configurado no Render.")

    base_url = app_public_url()
    if not base_url:
        raise RuntimeError("APP_PUBLIC_URL não configurado no Render.")

    event = query_db("SELECT * FROM billing_events WHERE id = ?", (event_id,), one=True)
    amount = plan_price(plan_key)
    payer_email = (company.get("email") if hasattr(company, "get") else company["email"]) or "cliente@email.com"
    payer_name = (company.get("responsible_name") if hasattr(company, "get") else company["responsible_name"]) or "Cliente"
    external_reference = event["external_reference"]

    payload = {
        "items": [
            {
                "id": f"pontofacil-{plan_key.lower()}",
                "title": f"{SYSTEM_COMMERCIAL_NAME} - Plano {plan_label(plan_key)}",
                "quantity": 1,
                "currency_id": "BRL",
                "unit_price": float(amount),
                "description": "Assinatura mensal do sistema de ponto online",
            }
        ],
        "payer": {
            "email": payer_email,
            "name": payer_name,
        },
        "external_reference": external_reference,
        "notification_url": f"{base_url}/webhook/mercadopago",
        "back_urls": {
            "success": f"{base_url}/pagamento/sucesso",
            "failure": f"{base_url}/pagamento/falha",
            "pending": f"{base_url}/pagamento/pendente",
        },
        "auto_return": "approved",
        "statement_descriptor": "PONTOFACILPRO",
        "payment_methods": {
            "installments": 12,
            "excluded_payment_types": [
                {"id": "ticket"}
            ],
        },
    }

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    response = requests.post(
        "https://api.mercadopago.com/checkout/preferences",
        json=payload,
        headers=headers,
        timeout=30,
    )

    try:
        data = response.json()
    except Exception:
        raise RuntimeError(f"Mercado Pago retornou resposta inválida: HTTP {response.status_code} - {response.text[:300]}")

    if response.status_code >= 400:
        message = data.get("message") or data.get("error") or str(data)
        raise RuntimeError(
            "Mercado Pago retornou erro ao criar Checkout Pro: "
            f"{message}. Confira se MP_ACCESS_TOKEN é o Access Token de produção copiado sem espaços."
        )

    if not (data.get("init_point") or data.get("sandbox_init_point")):
        raise RuntimeError(f"Mercado Pago criou a preferência, mas não retornou link de checkout: {data}")

    return data


def fetch_mercadopago_payment(payment_id: str) -> Dict[str, Any]:
    token = mp_access_token()
    if not token:
        raise RuntimeError("MP_ACCESS_TOKEN não configurado.")
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(f"https://api.mercadopago.com/v1/payments/{payment_id}", headers=headers, timeout=30)
    data = response.json()
    if response.status_code >= 400:
        raise RuntimeError(f"Erro ao consultar pagamento Mercado Pago: {data}")
    return data


def parse_external_reference(reference: str) -> Dict[str, str]:
    parts = (reference or "").split(":")
    result = {}
    for i in range(0, len(parts) - 1, 2):
        result[parts[i]] = parts[i + 1]
    return result


def activate_company_payment(company_id: int, plan_key: str, payment_id: str, event_id: Optional[int] = None) -> None:
    paid_until = (date.today() + timedelta(days=30)).strftime("%Y-%m-%d")
    execute_db(
        """
        UPDATE companies
        SET plan = ?, status = 'ATIVO', payment_status = 'PAGO', paid_until = ?, last_payment_at = ?
        WHERE id = ?
        """,
        (plan_key.upper(), paid_until, now_iso(), company_id),
    )
    if event_id:
        execute_db(
            """
            UPDATE billing_events
            SET status = 'PAGO', mp_payment_id = ?, paid_at = ?
            WHERE id = ?
            """,
            (str(payment_id), now_iso(), event_id),
        )


def block_overdue_companies() -> None:
    ensure_billing_ready()
    today = today_str()
    companies = query_db("SELECT * FROM companies WHERE status = 'ATIVO'")
    for company in companies:
        paid_until = company.get("paid_until") if hasattr(company, "get") else company["paid_until"]
        trial_until = company.get("trial_until") if hasattr(company, "get") else company["trial_until"]
        if paid_until and paid_until < today:
            execute_db("UPDATE companies SET status = 'BLOQUEADO', payment_status = 'VENCIDO' WHERE id = ?", (company["id"],))
        elif not paid_until and trial_until and trial_until < today:
            execute_db("UPDATE companies SET status = 'BLOQUEADO', payment_status = 'TESTE_EXPIRADO' WHERE id = ?", (company["id"],))


@app.route("/assinatura/pagar/<plan_key>", methods=["POST"])
@admin_required
def billing_create_payment(plan_key: str):
    ensure_billing_ready()
    company = current_company()
    if not company:
        flash("Nenhuma empresa ativa para gerar cobrança.", "danger")
        return redirect(url_for("subscription_page"))

    plan_key = (plan_key or "BASICO").upper()
    if plan_key not in PLAN_LIMITS:
        flash("Plano inválido.", "danger")
        return redirect(url_for("subscription_page"))

    try:
        event_id = create_billing_event(company["id"], plan_key)
        preference = create_mercadopago_preference(company, plan_key, event_id)
        checkout_url = preference.get("init_point") or preference.get("sandbox_init_point")
        execute_db(
            """
            UPDATE billing_events
            SET mp_preference_id = ?, checkout_url = ?
            WHERE id = ?
            """,
            (preference.get("id", ""), checkout_url or "", event_id),
        )
        if checkout_url:
            return redirect(checkout_url)
        flash("Cobrança criada, mas o Mercado Pago não retornou link de checkout.", "danger")
    except Exception as exc:
        flash(f"Erro ao gerar cobrança: {exc}", "danger")
    return redirect(url_for("subscription_page"))


@app.route("/webhook", methods=["GET", "POST"])
@app.route("/webhook/mercadopago", methods=["GET", "POST"])
def mercadopago_webhook():
    ensure_billing_ready()
    payload = request.get_json(silent=True) or {}
    topic = request.args.get("topic") or payload.get("type") or payload.get("topic") or payload.get("action")
    payment_id = request.args.get("id")

    if not payment_id and isinstance(payload.get("data"), dict):
        payment_id = payload["data"].get("id")
    if topic and "payment" not in str(topic):
        return {"status": "ignored", "topic": topic}, 200
    if not payment_id:
        return {"status": "ignored", "reason": "missing payment id"}, 200

    try:
        payment = fetch_mercadopago_payment(str(payment_id))
        status = payment.get("status", "")
        external_reference = payment.get("external_reference", "")
        ref = parse_external_reference(external_reference)
        company_id = int(ref.get("company", "0") or 0)
        plan_key = ref.get("plan", "BASICO")
        event_id = int(ref.get("event", "0") or 0)
        if status == "approved" and company_id:
            activate_company_payment(company_id, plan_key, str(payment_id), event_id or None)
        elif event_id:
            execute_db("UPDATE billing_events SET status = ?, mp_payment_id = ? WHERE id = ?", (status.upper() or "PENDENTE", str(payment_id), event_id))
        return {"status": "ok"}, 200
    except Exception as exc:
        print("Erro webhook Mercado Pago:", exc)
        return {"status": "error", "message": str(exc)}, 200


@app.route("/pagamento/sucesso")
def payment_success():
    flash("Pagamento recebido ou em processamento. A ativação automática ocorre após confirmação do Mercado Pago.", "success")
    return redirect(url_for("subscription_page"))


@app.route("/pagamento/falha")
def payment_failure():
    flash("Pagamento não aprovado. Tente novamente ou escolha outro meio de pagamento.", "danger")
    return redirect(url_for("subscription_page"))


@app.route("/pagamento/pendente")
def payment_pending():
    flash("Pagamento pendente. Assim que o Mercado Pago confirmar, o plano será ativado automaticamente.", "warning")
    return redirect(url_for("subscription_page"))




# =========================================================
# Automação SaaS - cobrança, bloqueio, liberação e avisos
# =========================================================
_SAAS_AUTOMATION_READY = False


def _safe_get(row, key, default=""):
    if not row:
        return default
    try:
        return row.get(key, default)
    except Exception:
        try:
            return row[key]
        except Exception:
            return default


def company_access_allowed(company) -> bool:
    """Regra central: ativo, pago ou teste válido libera acesso."""
    if not company:
        return False
    status = str(_safe_get(company, "status", "")).upper()
    payment_status = str(_safe_get(company, "payment_status", "")).upper()
    paid_until = str(_safe_get(company, "paid_until", "") or "")
    trial_until = str(_safe_get(company, "trial_until", "") or "")
    today = today_str()

    if status == "ATIVO" and (not paid_until or paid_until >= today):
        return True
    if payment_status == "PAGO" and paid_until and paid_until >= today:
        return True
    if status == "TESTE" and trial_until and trial_until >= today:
        return True
    return False


def ensure_saas_automation_ready() -> None:
    """Garante colunas para automação sem apagar dados."""
    global _SAAS_AUTOMATION_READY
    if _SAAS_AUTOMATION_READY:
        return
    try:
        ensure_billing_ready()
    except Exception:
        try:
            ensure_multiempresa_ready()
        except Exception:
            pass

    for column in ["blocked_at", "last_billing_alert_at", "next_billing_at"]:
        try:
            if not _column_exists("companies", column):
                execute_db(f"ALTER TABLE companies ADD COLUMN {column} TEXT")
        except Exception as exc:
            print(f"Aviso SaaS: não foi possível criar companies.{column}: {exc}")

    for column in ["notified_at", "webhook_payload"]:
        try:
            if not _column_exists("billing_events", column):
                execute_db(f"ALTER TABLE billing_events ADD COLUMN {column} TEXT")
        except Exception as exc:
            print(f"Aviso SaaS: não foi possível criar billing_events.{column}: {exc}")

    _SAAS_AUTOMATION_READY = True


@app.before_request
def saas_automation_bootstrap():
    if request.endpoint not in ("static", "health", "service_worker"):
        try:
            ensure_saas_automation_ready()
            auto_block_overdue_companies()
        except Exception as exc:
            print("Aviso automação SaaS:", exc)


def auto_block_overdue_companies() -> int:
    """Bloqueia empresas vencidas automaticamente, exceto super admin/Pet & Gatô interno."""
    today = today_str()
    changed = 0
    rows = query_db("SELECT * FROM companies")
    for company in rows:
        cid = _safe_get(company, "id")
        slug = str(_safe_get(company, "slug", ""))
        status = str(_safe_get(company, "status", "")).upper()
        paid_until = str(_safe_get(company, "paid_until", "") or "")
        trial_until = str(_safe_get(company, "trial_until", "") or "")

        # Mantém sua empresa base liberada, se desejar administrar sem cobrança.
        if slug == "pet-gato":
            continue

        if status == "ATIVO" and paid_until and paid_until < today:
            execute_db(
                "UPDATE companies SET status='BLOQUEADO', payment_status='VENCIDO', blocked_at=? WHERE id=?",
                (now_iso(), cid),
            )
            changed += 1
        elif status == "TESTE" and trial_until and trial_until < today:
            execute_db(
                "UPDATE companies SET status='BLOQUEADO', payment_status='TESTE_EXPIRADO', blocked_at=? WHERE id=?",
                (now_iso(), cid),
            )
            changed += 1
    return changed


def activate_company_payment(company_id: int, plan_key: str, payment_id: str, event_id: Optional[int] = None) -> None:
    """Libera empresa por 30 dias após pagamento aprovado."""
    paid_until = (date.today() + timedelta(days=30)).strftime("%Y-%m-%d")
    next_billing = (date.today() + timedelta(days=27)).strftime("%Y-%m-%d")
    execute_db(
        """
        UPDATE companies
        SET plan = ?, status = 'ATIVO', payment_status = 'PAGO', paid_until = ?,
            last_payment_at = ?, next_billing_at = ?, blocked_at = NULL
        WHERE id = ?
        """,
        (plan_key.upper(), paid_until, now_iso(), next_billing, company_id),
    )
    if event_id:
        execute_db(
            """
            UPDATE billing_events
            SET status = 'PAGO', mp_payment_id = ?, paid_at = ?
            WHERE id = ?
            """,
            (str(payment_id), now_iso(), event_id),
        )
    try:
        send_billing_whatsapp_notice(company_id, f"✅ Pagamento aprovado! Seu acesso ao PontoFácil Pro foi liberado até {paid_until}.")
    except Exception as exc:
        print("Aviso WhatsApp pagamento aprovado:", exc)


def send_billing_whatsapp_notice(company_id: int, message: str) -> Optional[str]:
    company = query_db("SELECT * FROM companies WHERE id = ?", (company_id,), one=True)
    phone = _safe_get(company, "phone", "")
    if not phone:
        return None
    link = build_whatsapp_link(phone, message)
    try:
        execute_db(
            "INSERT INTO system_logs (user_name, action, details, created_at) VALUES (?, ?, ?, ?)",
            ("Sistema", "WHATSAPP_COBRANCA", f"Empresa {company_id}: {message}", now_iso()),
        )
    except Exception:
        pass
    return link


def billing_reminder_message(company) -> str:
    name = _safe_get(company, "name", "Cliente")
    plan = plan_label(_safe_get(company, "plan", "BASICO")) if "plan_label" in globals() else _safe_get(company, "plan", "BASICO")
    paid_until = _safe_get(company, "paid_until", "-") or "-"
    return (
        f"Olá, {name}!%0A%0A"
        f"Sua assinatura do PontoFácil Pro está pendente ou próxima do vencimento.%0A"
        f"Plano: {plan}%0A"
        f"Pago até: {paid_until}%0A%0A"
        f"Acesse o painel e clique em Assinatura para regularizar.%0A"
        f"{app_public_url()}/assinatura"
    )


@app.route("/financeiro-saas")
@admin_required
def financeiro_saas():
    if not is_super_admin():
        flash("Acesso restrito ao administrador master.", "danger")
        return redirect(url_for("dashboard"))
    ensure_saas_automation_ready()
    rows = query_db("SELECT * FROM companies ORDER BY id ASC")
    companies = []
    total_mrr = 0.0
    for c in rows:
        row = dict(c)
        plan = str(row.get("plan") or "BASICO").upper()
        row["price"] = plan_price(plan) if "plan_price" in globals() else 0
        row["access_ok"] = company_access_allowed(row)
        row["billing_whatsapp"] = build_whatsapp_link(row.get("phone", ""), billing_reminder_message(row))
        if row.get("status") == "ATIVO":
            total_mrr += float(row["price"])
        companies.append(row)
    events = query_db("SELECT * FROM billing_events ORDER BY created_at DESC, id DESC LIMIT 100")
    return render_template("financeiro_saas.html", companies=companies, events=events, total_mrr=total_mrr)


@app.route("/financeiro-saas/verificar-pagamentos", methods=["POST"])
@admin_required
def financeiro_verificar_pagamentos():
    if not is_super_admin():
        flash("Acesso restrito.", "danger")
        return redirect(url_for("dashboard"))
    ensure_saas_automation_ready()
    pending = query_db("SELECT * FROM billing_events WHERE status IN ('AGUARDANDO_PAGAMENTO','PENDENTE') ORDER BY id DESC LIMIT 50")
    checked = 0
    approved = 0
    for event in pending:
        payment_id = _safe_get(event, "mp_payment_id", "")
        if not payment_id:
            continue
        try:
            payment = fetch_mercadopago_payment(str(payment_id))
            checked += 1
            if payment.get("status") == "approved":
                ref = parse_external_reference(payment.get("external_reference", ""))
                activate_company_payment(int(ref.get("company", 0)), ref.get("plan", "BASICO"), str(payment_id), int(_safe_get(event, "id")))
                approved += 1
        except Exception as exc:
            print("Erro verificação manual pagamento:", exc)
    flash(f"Verificação concluída. Consultados: {checked}. Aprovados: {approved}.", "success")
    return redirect(url_for("financeiro_saas"))

@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.json

    if data.get("type") == "payment":
        payment_id = data["data"]["id"]

        url = f"https://api.mercadopago.com/v1/payments/{payment_id}"
        headers = {
            "Authorization": f"Bearer {os.getenv('MP_ACCESS_TOKEN')}"
        }

        resp = requests.get(url, headers=headers).json()

        if resp.get("status") == "approved":
            external_ref = resp.get("external_reference")

            # 🔥 ATIVA A EMPRESA AQUI
            liberar_empresa(external_ref)

    return "ok"

@app.route("/saas/status")
def saas_status():
    return {"status": "ok", "system": "PontoFácil Pro", "time": now_iso()}, 200

# =========================================================
# Fim Automação SaaS
# =========================================================


# Inicializa as tabelas tanto no Render/PostgreSQL quanto no Windows/SQLite.
with app.app_context():
    init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, debug=True)
