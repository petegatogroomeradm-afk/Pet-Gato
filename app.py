from __future__ import annotations

import os
import shutil
import sqlite3
import psycopg2
from psycopg2.extras import RealDictCursor
from urllib.parse import urlparse
from datetime import datetime, date, timedelta
from functools import wraps
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

from flask import (
    Flask,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
DATABASE = INSTANCE_DIR / "relogio_ponto.db"
EXPORTS_DIR = BASE_DIR / "exports"
BACKUPS_DIR = BASE_DIR / "backups"

WEEKDAY_LABELS = {
    0: "Segunda-feira",
    1: "Terça-feira",
    2: "Quarta-feira",
    3: "Quinta-feira",
    4: "Sexta-feira",
    5: "Sábado",
    6: "Domingo",
}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "petegato_relogio_2026")
app.config["DATABASE"] = str(DATABASE)
app.config["EXPORTS_DIR"] = str(EXPORTS_DIR)
app.config["BACKUPS_DIR"] = str(BACKUPS_DIR)

STORE_ALLOWED_IP = os.environ.get("STORE_ALLOWED_IP", "").strip()
ADMIN_SECRET_KEY = os.environ.get("ADMIN_SECRET_KEY", "").strip()


def get_client_ip() -> str:
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or ""


def is_store_network() -> bool:
    client_ip = get_client_ip()
    return bool(STORE_ALLOWED_IP and client_ip == STORE_ALLOWED_IP)


def has_admin_key() -> bool:
    key = request.args.get("chave", "").strip()
    if key and ADMIN_SECRET_KEY and key == ADMIN_SECRET_KEY:
        session["admin_key_ok"] = True
        return True
    return bool(session.get("admin_key_ok"))


@app.before_request
def security_gate():
    allowed_public = ["static", "service_worker", "health"]

    if request.endpoint in allowed_public:
        return None

    if request.path.startswith("/ponto") or request.path == "/":
        if not is_store_network() and not has_admin_key():
            return render_template(
                "blocked.html",
                ip=get_client_ip()
            ), 403

    admin_paths = [
        "/login",
        "/dashboard",
        "/funcionarios",
        "/registros",
        "/relatorios",
        "/jornadas",
        "/configuracoes",
        "/logout",
    ]

    if any(request.path.startswith(path) for path in admin_paths):
        if not has_admin_key():
            return render_template(
                "blocked.html",
                ip=get_client_ip()
            ), 403

    return None

# ---------------------------
# Database helpers
# ---------------------------
def get_db():
    if "db" not in g:
        database_url = os.getenv("DATABASE_URL")

        if database_url:
            g.db = psycopg2.connect(database_url, cursor_factory=RealDictCursor)
        else:
            g.db = sqlite3.connect(app.config["DATABASE"])
            g.db.row_factory = sqlite3.Row

    return g.db


@app.teardown_appcontext
def close_db(exc: Optional[BaseException]) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


# ---------------------------
# Utils
# ---------------------------
def now_dt() -> datetime:
    return datetime.now()


def now_iso() -> str:
    return now_dt().strftime("%Y-%m-%d %H:%M:%S")


def today_str() -> str:
    return date.today().strftime("%Y-%m-%d")


def parse_dt(value: str) -> datetime:
    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_hhmm(value: str) -> Tuple[int, int]:
    hh, mm = value.split(":")
    return int(hh), int(mm)


def combine_day_time(day: date, hhmm: str) -> datetime:
    hh, mm = parse_hhmm(hhmm)
    return datetime(day.year, day.month, day.day, hh, mm)


def format_dt(value: Optional[str]) -> str:
    if not value:
        return "-"
    return parse_dt(value).strftime("%d/%m/%Y %H:%M:%S")


def format_minutes(total_minutes: int) -> str:
    sign = "-" if total_minutes < 0 else ""
    total_minutes = abs(total_minutes)
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{sign}{hours:02d}:{minutes:02d}"


def record_sequence() -> List[str]:
    return ["ENTRADA", "SAIDA_INTERVALO", "RETORNO_INTERVALO", "SAIDA"]


def normalize_phone(phone: str) -> str:
    return "".join(ch for ch in phone if ch.isdigit())


def build_whatsapp_link(phone: str, message: str) -> Optional[str]:
    digits = normalize_phone(phone)
    if not digits:
        return None
    return f"https://wa.me/{digits}?text={quote(message)}"


def weekday_name(idx: int) -> str:
    return WEEKDAY_LABELS.get(idx, str(idx))


def adapt_query(query: str) -> str:
    if os.getenv("DATABASE_URL"):
        return query.replace("?", "%s")
    return query


def query_db(query: str, params: tuple = (), one: bool = False) -> Any:
    cur = get_db().cursor()
    cur.execute(adapt_query(query), params)
    rows = cur.fetchall()
    cur.close()
    return (rows[0] if rows else None) if one else rows


def execute_db(query: str, params: tuple = ()) -> int:
    db = get_db()
    cur = db.cursor()
    cur.execute(adapt_query(query), params)
    db.commit()

    lastrowid = 0
    try:
        lastrowid = cur.lastrowid
    except Exception:
        pass

    cur.close()
    return lastrowid


def get_setting(key: str, default: str = "") -> str:
    row = query_db("SELECT value FROM settings WHERE key = ?", (key,), one=True)
    return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    execute_db(
        """
        INSERT INTO settings (key, value, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
        """,
        (key, value, now_iso()),
    )


def init_db() -> None:
    INSTANCE_DIR.mkdir(parents=True, exist_ok=True)
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)

    schema = """
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'admin',
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS schedules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        description TEXT NOT NULL,
        entry_time TEXT NOT NULL,
        lunch_out_time TEXT NOT NULL,
        lunch_in_time TEXT NOT NULL,
        exit_time TEXT NOT NULL,
        tolerance_minutes INTEGER NOT NULL DEFAULT 10,
        active INTEGER NOT NULL DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        cpf TEXT,
        registration TEXT NOT NULL UNIQUE,
        role_name TEXT,
        phone TEXT,
        admission_date TEXT,
        pin_hash TEXT NOT NULL,
        schedule_id INTEGER,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL,
        FOREIGN KEY (schedule_id) REFERENCES schedules(id)
    );

    CREATE TABLE IF NOT EXISTS employee_schedule_days (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        weekday INTEGER NOT NULL,
        schedule_id INTEGER NOT NULL,
        UNIQUE(employee_id, weekday),
        FOREIGN KEY (employee_id) REFERENCES employees(id),
        FOREIGN KEY (schedule_id) REFERENCES schedules(id)
    );

    CREATE TABLE IF NOT EXISTS time_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        record_type TEXT NOT NULL,
        record_time TEXT NOT NULL,
        notes TEXT,
        ip_origin TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY (employee_id) REFERENCES employees(id)
    );

    CREATE TABLE IF NOT EXISTS time_adjustments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_id INTEGER,
        employee_id INTEGER NOT NULL,
        admin_user_id INTEGER NOT NULL,
        old_value TEXT,
        new_value TEXT NOT NULL,
        reason TEXT NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY (record_id) REFERENCES time_records(id),
        FOREIGN KEY (employee_id) REFERENCES employees(id),
        FOREIGN KEY (admin_user_id) REFERENCES users(id)
    );

    CREATE TABLE IF NOT EXISTS system_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_name TEXT,
        action TEXT NOT NULL,
        details TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL,
        updated_at TEXT NOT NULL
    );
    """

    with sqlite3.connect(app.config["DATABASE"]) as conn:
        conn.executescript(schema)
        conn.commit()

    with sqlite3.connect(app.config["DATABASE"]) as conn:
        conn.row_factory = sqlite3.Row
        schedule = conn.execute("SELECT id FROM schedules LIMIT 1").fetchone()
        if not schedule:
            conn.execute(
                """
                INSERT INTO schedules (
                    description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes, active
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                ("Jornada padrão 08h às 17h", "08:00", "12:00", "13:00", "17:00", 10, 1),
            )
            conn.execute(
                """
                INSERT INTO schedules (
                    description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes, active
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                ("Terça a Sexta 09h às 18h", "09:00", "12:00", "13:00", "18:00", 10, 1),
            )
            conn.execute(
                """
                INSERT INTO schedules (
                    description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes, active
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                ("Sábado 08h às 17h", "08:00", "12:00", "13:00", "17:00", 10, 1),
            )

        admin = conn.execute("SELECT id FROM users WHERE username = ?", ("admin",)).fetchone()
        if not admin:
            conn.execute(
                "INSERT INTO users (name, username, password_hash, role, active, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    "Administrador",
                    "admin",
                    generate_password_hash("Pet&gato3264"),
                    "admin",
                    1,
                    now_iso(),
                ),
            )

        existing = conn.execute("SELECT key FROM settings WHERE key = 'company_whatsapp'").fetchone()
        if not existing:
            conn.execute(
                "INSERT INTO settings (key, value, updated_at) VALUES (?, ?, ?)",
                ("company_whatsapp", "", now_iso()),
            )
        conn.commit()


def get_schedule_for_day(employee_id: int, ref_date: str) -> Optional[sqlite3.Row]:
    ref = parse_date(ref_date)
    weekday = ref.weekday()
    schedule = query_db(
        """
        SELECT s.*
        FROM employee_schedule_days esd
        JOIN schedules s ON s.id = esd.schedule_id
        WHERE esd.employee_id = ? AND esd.weekday = ? AND s.active = 1
        """,
        (employee_id, weekday),
        one=True,
    )
    if schedule:
        return schedule

    return query_db(
        "SELECT s.* FROM employees e LEFT JOIN schedules s ON s.id = e.schedule_id WHERE e.id = ?",
        (employee_id,),
        one=True,
    )


def get_employee_week_schedule_map(employee_id: int) -> Dict[int, sqlite3.Row]:
    rows = query_db(
        """
        SELECT esd.weekday, s.id AS schedule_id, s.description
        FROM employee_schedule_days esd
        JOIN schedules s ON s.id = esd.schedule_id
        WHERE esd.employee_id = ?
        ORDER BY esd.weekday
        """,
        (employee_id,),
    )
    return {row["weekday"]: row for row in rows}


def set_employee_week_schedule(employee_id: int, weekday: int, schedule_id: Optional[int]) -> None:
    execute_db("DELETE FROM employee_schedule_days WHERE employee_id = ? AND weekday = ?", (employee_id, weekday))
    if schedule_id:
        execute_db(
            "INSERT INTO employee_schedule_days (employee_id, weekday, schedule_id) VALUES (?, ?, ?)",
            (employee_id, weekday, schedule_id),
        )


def next_record_type(employee_id: int, ref_date: Optional[str] = None) -> str:
    ref_date = ref_date or today_str()
    records = query_db(
        """
        SELECT record_type FROM time_records
        WHERE employee_id = ? AND date(record_time) = ?
        ORDER BY record_time ASC, id ASC
        """,
        (employee_id, ref_date),
    )
    index = len(records)
    sequence = record_sequence()
    if index >= len(sequence):
        return "COMPLETO"
    return sequence[index]


def calculate_day_summary(employee_id: int, ref_date: str) -> Dict[str, Any]:
    employee = query_db(
        """
        SELECT e.*
        FROM employees e
        WHERE e.id = ?
        """,
        (employee_id,),
        one=True,
    )
    records = query_db(
        """
        SELECT * FROM time_records
        WHERE employee_id = ? AND date(record_time) = ?
        ORDER BY record_time ASC, id ASC
        """,
        (employee_id, ref_date),
    )
    schedule = get_schedule_for_day(employee_id, ref_date)
    sequence = {row["record_type"]: row["record_time"] for row in records}
    ref_day = parse_date(ref_date)

    scheduled_entry = combine_day_time(ref_day, schedule["entry_time"]) if schedule and schedule["entry_time"] else None
    scheduled_lunch_out = combine_day_time(ref_day, schedule["lunch_out_time"]) if schedule and schedule["lunch_out_time"] else None
    scheduled_lunch_in = combine_day_time(ref_day, schedule["lunch_in_time"]) if schedule and schedule["lunch_in_time"] else None
    scheduled_exit = combine_day_time(ref_day, schedule["exit_time"]) if schedule and schedule["exit_time"] else None

    worked_minutes = 0
    if sequence.get("ENTRADA") and sequence.get("SAIDA"):
        start = parse_dt(sequence["ENTRADA"])
        end = parse_dt(sequence["SAIDA"])
        worked_minutes = int((end - start).total_seconds() // 60)
        if sequence.get("SAIDA_INTERVALO") and sequence.get("RETORNO_INTERVALO"):
            lunch_out = parse_dt(sequence["SAIDA_INTERVALO"])
            lunch_in = parse_dt(sequence["RETORNO_INTERVALO"])
            worked_minutes -= int((lunch_in - lunch_out).total_seconds() // 60)

    scheduled_worked = 0
    if scheduled_entry and scheduled_exit and scheduled_lunch_out and scheduled_lunch_in:
        scheduled_worked = int((scheduled_exit - scheduled_entry).total_seconds() // 60)
        scheduled_worked -= int((scheduled_lunch_in - scheduled_lunch_out).total_seconds() // 60)

    delay_minutes = 0
    early_exit_minutes = 0
    extra_minutes = 0
    lunch_less_than_expected = False
    inconsistency = False

    if sequence.get("ENTRADA") and scheduled_entry:
        actual_entry = parse_dt(sequence["ENTRADA"])
        tolerance = schedule["tolerance_minutes"] if schedule else 0
        diff = int((actual_entry - scheduled_entry).total_seconds() // 60)
        if diff > tolerance:
            delay_minutes = diff

    if sequence.get("SAIDA") and scheduled_exit:
        actual_exit = parse_dt(sequence["SAIDA"])
        diff = int((scheduled_exit - actual_exit).total_seconds() // 60)
        if diff > 0:
            early_exit_minutes = diff

    if worked_minutes and scheduled_worked:
        extra_minutes = worked_minutes - scheduled_worked

    if sequence.get("SAIDA_INTERVALO") and sequence.get("RETORNO_INTERVALO") and scheduled_lunch_out and scheduled_lunch_in:
        actual_lunch = int((parse_dt(sequence["RETORNO_INTERVALO"]) - parse_dt(sequence["SAIDA_INTERVALO"])).total_seconds() // 60)
        expected_lunch = int((scheduled_lunch_in - scheduled_lunch_out).total_seconds() // 60)
        lunch_less_than_expected = actual_lunch < expected_lunch

    if len(records) not in (0, 4):
        inconsistency = True

    return {
        "employee": employee,
        "records": records,
        "sequence": sequence,
        "schedule": schedule,
        "worked_minutes": worked_minutes,
        "scheduled_minutes": scheduled_worked,
        "delay_minutes": delay_minutes,
        "early_exit_minutes": early_exit_minutes,
        "extra_minutes": extra_minutes,
        "lunch_less_than_expected": lunch_less_than_expected,
        "inconsistency": inconsistency,
        "next_type": next_record_type(employee_id, ref_date),
        "weekday_name": weekday_name(ref_day.weekday()),
    }


def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            flash("Faça login para acessar o painel.", "warning")
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper


def log_action(action: str, details: str = "") -> None:
    username = session.get("user_name", "Sistema")
    execute_db(
        "INSERT INTO system_logs (user_name, action, details, created_at) VALUES (?, ?, ?, ?)",
        (username, action, details, now_iso()),
    )


def month_bounds(year: int, month: int) -> Tuple[str, str]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def generate_excel_report(year: int, month: int) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Espelho de Ponto"

    headers = [
        "Funcionário", "Matrícula", "Data", "Dia", "Jornada", "Entrada", "Saída Intervalo",
        "Retorno Intervalo", "Saída", "Horas Trabalhadas", "Atraso", "Saída Antecipada",
        "Saldo (Extra/Negativo)", "Inconsistência"
    ]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    start, end = month_bounds(year, month)
    employees = query_db("SELECT * FROM employees WHERE active = 1 ORDER BY name")

    for emp in employees:
        current = parse_date(start)
        end_date = parse_date(end)
        while current < end_date:
            summary = calculate_day_summary(emp["id"], current.strftime("%Y-%m-%d"))
            seq = summary["sequence"]
            ws.append([
                emp["name"],
                emp["registration"],
                current.strftime("%d/%m/%Y"),
                summary["weekday_name"],
                summary["schedule"]["description"] if summary["schedule"] else "Sem jornada",
                parse_dt(seq["ENTRADA"]).strftime("%H:%M:%S") if seq.get("ENTRADA") else "",
                parse_dt(seq["SAIDA_INTERVALO"]).strftime("%H:%M:%S") if seq.get("SAIDA_INTERVALO") else "",
                parse_dt(seq["RETORNO_INTERVALO"]).strftime("%H:%M:%S") if seq.get("RETORNO_INTERVALO") else "",
                parse_dt(seq["SAIDA"]).strftime("%H:%M:%S") if seq.get("SAIDA") else "",
                format_minutes(summary["worked_minutes"]),
                format_minutes(summary["delay_minutes"]),
                format_minutes(summary["early_exit_minutes"]),
                format_minutes(summary["extra_minutes"]),
                "SIM" if summary["inconsistency"] else "NÃO",
            ])
            current += timedelta(days=1)

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[column].width = min(max_length + 2, 30)

    file_path = EXPORTS_DIR / f"relatorio_ponto_{year}_{month:02d}.xlsx"
    wb.save(file_path)
    return file_path


def generate_pdf_report(employee_id: int, year: int, month: int) -> Path:
    employee = query_db("SELECT * FROM employees WHERE id = ?", (employee_id,), one=True)
    file_path = EXPORTS_DIR / f"espelho_ponto_{employee['registration']}_{year}_{month:02d}.pdf"

    doc = SimpleDocTemplate(str(file_path), pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Espelho de Ponto - Pet & Gatô", styles["Title"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Funcionário: {employee['name']} | Matrícula: {employee['registration']}", styles["Normal"]))
    story.append(Paragraph(f"Referência: {month:02d}/{year}", styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [["Data", "Dia", "Jornada", "Entrada", "Saída Intervalo", "Retorno Intervalo", "Saída", "Horas", "Atraso", "Saldo", "Inconsistência"]]
    start, end = month_bounds(year, month)
    current = parse_date(start)
    end_date = parse_date(end)
    while current < end_date:
        summary = calculate_day_summary(employee_id, current.strftime("%Y-%m-%d"))
        seq = summary["sequence"]
        data.append([
            current.strftime("%d/%m/%Y"),
            summary["weekday_name"],
            summary["schedule"]["description"] if summary["schedule"] else "Sem jornada",
            parse_dt(seq["ENTRADA"]).strftime("%H:%M:%S") if seq.get("ENTRADA") else "",
            parse_dt(seq["SAIDA_INTERVALO"]).strftime("%H:%M:%S") if seq.get("SAIDA_INTERVALO") else "",
            parse_dt(seq["RETORNO_INTERVALO"]).strftime("%H:%M:%S") if seq.get("RETORNO_INTERVALO") else "",
            parse_dt(seq["SAIDA"]).strftime("%H:%M:%S") if seq.get("SAIDA") else "",
            format_minutes(summary["worked_minutes"]),
            format_minutes(summary["delay_minutes"]),
            format_minutes(summary["extra_minutes"]),
            "SIM" if summary["inconsistency"] else "NÃO",
        ])
        current += timedelta(days=1)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
    ]))
    story.append(table)
    story.append(Spacer(1, 25))
    story.append(Paragraph("Assinatura do colaborador: ____________________________________________", styles["Normal"]))

    doc.build(story)
    return file_path


def create_backup() -> Path:
    timestamp = now_dt().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUPS_DIR / f"backup_{timestamp}.db"
    if DATABASE.exists():
        shutil.copy2(DATABASE, backup_path)
    return backup_path


# ---------------------------
# Routes - auth
# ---------------------------
@app.route("/")
def index():
    return redirect(url_for("punch"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = query_db("SELECT * FROM users WHERE username = ? AND active = 1", (username,), one=True)
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["user_name"] = user["name"]
            log_action("LOGIN", f"Usuário {username} entrou no sistema")
            flash("Login realizado com sucesso.", "success")
            return redirect(url_for("dashboard"))
        flash("Usuário ou senha inválidos.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    if session.get("user_id"):
        log_action("LOGOUT", "Usuário saiu do sistema")
    session.clear()
    flash("Sessão encerrada.", "info")
    return redirect(url_for("login"))


# ---------------------------
# Routes - punch
# ---------------------------
@app.route("/ponto", methods=["GET", "POST"])
def punch():
    preview = None
    whatsapp_links = {}

    if request.method == "POST":
        registration = request.form.get("registration", "").strip()
        pin = request.form.get("pin", "").strip()

        employee = query_db("SELECT * FROM employees WHERE registration = ? AND active = 1", (registration,), one=True)
        if not employee:
            flash("Matrícula não encontrada.", "danger")
            return render_template("punch.html", preview=preview, current_time=now_dt(), whatsapp_links=whatsapp_links)

        if not check_password_hash(employee["pin_hash"], pin):
            flash("PIN inválido.", "danger")
            return render_template("punch.html", preview=preview, current_time=now_dt(), whatsapp_links=whatsapp_links)

        next_type = next_record_type(employee["id"])
        if next_type == "COMPLETO":
            flash("Todas as marcações do dia já foram registradas para este funcionário.", "warning")
            return render_template("punch.html", preview=preview, current_time=now_dt(), whatsapp_links=whatsapp_links)

        moment = now_iso()
        execute_db(
            """
            INSERT INTO time_records (employee_id, record_type, record_time, notes, ip_origin, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (employee["id"], next_type, moment, "Registro automático", request.remote_addr or "local", now_iso()),
        )
        preview = calculate_day_summary(employee["id"], today_str())
        log_action("REGISTRO_PONTO", f"{employee['name']} registrou {next_type}")
        flash(f"{next_type.replace('_', ' ').title()} registrada com sucesso para {employee['name']}.", "success")

        human_type = next_type.replace("_", " ").title()
        message = (
            f"Pet & Gatô - comprovante de ponto%0A"
            f"Funcionário: {employee['name']}%0A"
            f"Matrícula: {employee['registration']}%0A"
            f"Marcação: {human_type}%0A"
            f"Data/Hora: {datetime.strptime(moment, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M:%S')}"
        )
        employee_phone = employee["phone"] or ""
        manager_phone = get_setting("company_whatsapp", "")
        employee_link = build_whatsapp_link(employee_phone, message)
        manager_link = build_whatsapp_link(manager_phone, message)
        if employee_link:
            whatsapp_links["employee"] = employee_link
        if manager_link:
            whatsapp_links["manager"] = manager_link

    return render_template("punch.html", preview=preview, current_time=now_dt(), whatsapp_links=whatsapp_links)


# ---------------------------
# Routes - dashboard/admin
# ---------------------------
@app.route("/dashboard")
@admin_required
def dashboard():
    today = today_str()
    total_employees = query_db("SELECT COUNT(*) AS total FROM employees WHERE active = 1", one=True)["total"]
    present_rows = query_db("SELECT COUNT(DISTINCT employee_id) AS total FROM time_records WHERE date(record_time) = ?", (today,), one=True)
    present_today = present_rows["total"]
    recent_records = query_db(
        """
        SELECT tr.*, e.name AS employee_name, e.registration
        FROM time_records tr
        JOIN employees e ON e.id = tr.employee_id
        ORDER BY tr.record_time DESC
        LIMIT 10
        """
    )
    employees = query_db("SELECT id, name FROM employees WHERE active = 1 ORDER BY name")

    delayed_count = 0
    inconsistent_count = 0
    for employee in employees:
        summary = calculate_day_summary(employee["id"], today)
        if summary["delay_minutes"] > 0:
            delayed_count += 1
        if summary["inconsistency"]:
            inconsistent_count += 1

    return render_template(
        "dashboard.html",
        total_employees=total_employees,
        present_today=present_today,
        absent_today=max(total_employees - present_today, 0),
        delayed_count=delayed_count,
        inconsistent_count=inconsistent_count,
        recent_records=recent_records,
        format_dt=format_dt,
    )


@app.route("/funcionarios", methods=["GET", "POST"])
@admin_required
def employees():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        cpf = request.form.get("cpf", "").strip()
        registration = request.form.get("registration", "").strip()
        role_name = request.form.get("role_name", "").strip()
        phone = request.form.get("phone", "").strip()
        admission_date = request.form.get("admission_date", "").strip() or None
        pin = request.form.get("pin", "").strip()
        schedule_id = request.form.get("schedule_id", type=int)
        schedule_tue_fri = request.form.get("schedule_tue_fri", type=int)
        schedule_sat = request.form.get("schedule_sat", type=int)

        if not all([name, registration, pin]):
            flash("Nome, matrícula e PIN são obrigatórios.", "danger")
        else:
            try:
                employee_id = execute_db(
                    """
                    INSERT INTO employees (name, cpf, registration, role_name, phone, admission_date, pin_hash, schedule_id, active, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                    """,
                    (name, cpf, registration, role_name, phone, admission_date, generate_password_hash(pin), schedule_id, now_iso()),
                )
                for weekday in [1, 2, 3, 4]:
                    if schedule_tue_fri:
                        set_employee_week_schedule(employee_id, weekday, schedule_tue_fri)
                if schedule_sat:
                    set_employee_week_schedule(employee_id, 5, schedule_sat)
                log_action("CADASTRO_FUNCIONARIO", f"Funcionário {name} cadastrado")
                flash("Funcionário cadastrado com sucesso.", "success")
                return redirect(url_for("employees"))
            except sqlite3.IntegrityError:
                flash("A matrícula já existe. Use outra matrícula.", "danger")

    employee_list = query_db(
        """
        SELECT e.*, s.description AS schedule_description
        FROM employees e
        LEFT JOIN schedules s ON s.id = e.schedule_id
        ORDER BY e.name
        """
    )
    schedules = query_db("SELECT * FROM schedules WHERE active = 1 ORDER BY description")
    week_maps = {emp["id"]: get_employee_week_schedule_map(emp["id"]) for emp in employee_list}
    return render_template("employees.html", employees=employee_list, schedules=schedules, week_maps=week_maps, weekday_name=weekday_name)


@app.route("/funcionarios/<int:employee_id>/toggle")
@admin_required
def toggle_employee(employee_id: int):
    employee = query_db("SELECT * FROM employees WHERE id = ?", (employee_id,), one=True)
    if employee:
        new_active = 0 if employee["active"] else 1
        execute_db("UPDATE employees SET active = ? WHERE id = ?", (new_active, employee_id))
        log_action("ATUALIZACAO_FUNCIONARIO", f"Funcionário {employee['name']} ativo={new_active}")
        flash("Status do funcionário atualizado.", "success")
    return redirect(url_for("employees"))


@app.route("/funcionarios/<int:employee_id>/jornadas", methods=["POST"])
@admin_required
def update_employee_schedules(employee_id: int):
    employee = query_db("SELECT * FROM employees WHERE id = ?", (employee_id,), one=True)
    if not employee:
        flash("Funcionário não encontrado.", "danger")
        return redirect(url_for("employees"))

    default_schedule_id = request.form.get("schedule_id", type=int)
    tue_fri = request.form.get("schedule_tue_fri", type=int)
    saturday = request.form.get("schedule_sat", type=int)
    execute_db("UPDATE employees SET schedule_id = ? WHERE id = ?", (default_schedule_id, employee_id))
    for weekday in range(7):
        set_employee_week_schedule(employee_id, weekday, None)
    for weekday in [1, 2, 3, 4]:
        if tue_fri:
            set_employee_week_schedule(employee_id, weekday, tue_fri)
    if saturday:
        set_employee_week_schedule(employee_id, 5, saturday)

    log_action("ATUALIZACAO_JORNADA_FUNCIONARIO", f"Jornadas atualizadas para {employee['name']}")
    flash("Jornadas do funcionário atualizadas.", "success")
    return redirect(url_for("employees"))


@app.route("/registros")
@admin_required
def records():
    employee_id = request.args.get("employee_id", type=int)
    ref_date = request.args.get("ref_date", today_str())
    params = []
    where = ["date(tr.record_time) = ?"]
    params.append(ref_date)
    if employee_id:
        where.append("tr.employee_id = ?")
        params.append(employee_id)

    record_list = query_db(
        f"""
        SELECT tr.*, e.name AS employee_name, e.registration
        FROM time_records tr
        JOIN employees e ON e.id = tr.employee_id
        WHERE {' AND '.join(where)}
        ORDER BY tr.record_time ASC
        """,
        tuple(params),
    )
    employees = query_db("SELECT id, name FROM employees WHERE active = 1 ORDER BY name")
    return render_template("records.html", records=record_list, employees=employees, ref_date=ref_date, employee_id=employee_id, format_dt=format_dt)


@app.route("/registros/<int:record_id>/ajustar", methods=["POST"])
@admin_required
def adjust_record(record_id: int):
    record = query_db("SELECT * FROM time_records WHERE id = ?", (record_id,), one=True)
    if not record:
        flash("Registro não encontrado.", "danger")
        return redirect(url_for("records"))

    new_value = request.form.get("new_value", "").strip()
    reason = request.form.get("reason", "").strip()
    if not new_value or not reason:
        flash("Nova data/hora e motivo são obrigatórios.", "danger")
        return redirect(url_for("records", ref_date=parse_dt(record["record_time"]).strftime("%Y-%m-%d")))

    try:
        datetime.strptime(new_value, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        flash("Formato inválido. Use YYYY-MM-DD HH:MM:SS.", "danger")
        return redirect(url_for("records", ref_date=parse_dt(record["record_time"]).strftime("%Y-%m-%d")))

    execute_db(
        "INSERT INTO time_adjustments (record_id, employee_id, admin_user_id, old_value, new_value, reason, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (record_id, record["employee_id"], session["user_id"], record["record_time"], new_value, reason, now_iso()),
    )
    execute_db("UPDATE time_records SET record_time = ? WHERE id = ?", (new_value, record_id))
    log_action("AJUSTE_REGISTRO", f"Registro {record_id} ajustado. Motivo: {reason}")
    flash("Registro ajustado com sucesso.", "success")
    return redirect(url_for("records", ref_date=parse_dt(new_value).strftime("%Y-%m-%d"), employee_id=record["employee_id"]))


@app.route("/relatorios", methods=["GET", "POST"])
@admin_required
def reports():
    employees = query_db("SELECT id, name, registration FROM employees WHERE active = 1 ORDER BY name")
    today = date.today()

    if request.method == "POST":
        action = request.form.get("action")
        year = request.form.get("year", type=int) or today.year
        month = request.form.get("month", type=int) or today.month
        employee_id = request.form.get("employee_id", type=int)

        if action == "excel":
            file_path = generate_excel_report(year, month)
            log_action("EXPORTAR_EXCEL", f"Relatório mensal {month:02d}/{year}")
            return send_file(file_path, as_attachment=True)

        if action == "pdf":
            if not employee_id:
                flash("Selecione um funcionário para gerar o PDF.", "danger")
            else:
                file_path = generate_pdf_report(employee_id, year, month)
                log_action("EXPORTAR_PDF", f"Espelho de ponto do funcionário {employee_id} - {month:02d}/{year}")
                return send_file(file_path, as_attachment=True)

        if action == "backup":
            backup_path = create_backup()
            log_action("BACKUP", f"Backup gerado: {backup_path.name}")
            return send_file(backup_path, as_attachment=True)

    return render_template("reports.html", employees=employees, today=today)


@app.route("/jornadas", methods=["GET", "POST"])
@admin_required
def schedules():
    if request.method == "POST":
        description = request.form.get("description", "").strip()
        entry_time = request.form.get("entry_time", "").strip()
        lunch_out_time = request.form.get("lunch_out_time", "").strip()
        lunch_in_time = request.form.get("lunch_in_time", "").strip()
        exit_time = request.form.get("exit_time", "").strip()
        tolerance_minutes = request.form.get("tolerance_minutes", type=int) or 10

        if all([description, entry_time, lunch_out_time, lunch_in_time, exit_time]):
            execute_db(
                """
                INSERT INTO schedules (description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes, active)
                VALUES (?, ?, ?, ?, ?, ?, 1)
                """,
                (description, entry_time, lunch_out_time, lunch_in_time, exit_time, tolerance_minutes),
            )
            log_action("CADASTRO_JORNADA", f"Jornada cadastrada: {description}")
            flash("Jornada cadastrada com sucesso.", "success")
            return redirect(url_for("schedules"))
        flash("Preencha todos os campos da jornada.", "danger")

    items = query_db("SELECT * FROM schedules ORDER BY id DESC")
    return render_template("schedules.html", schedules=items)


@app.route("/configuracoes", methods=["GET", "POST"])
@admin_required
def settings_page():
    if request.method == "POST":
        company_whatsapp = request.form.get("company_whatsapp", "").strip()
        set_setting("company_whatsapp", company_whatsapp)
        log_action("CONFIGURACOES", "WhatsApp da empresa atualizado")
        flash("Configurações salvas com sucesso.", "success")
        return redirect(url_for("settings_page"))

    return render_template("settings.html", company_whatsapp=get_setting("company_whatsapp", ""))


@app.route("/health")
def health():
    return {"status": "ok", "time": now_iso()}


@app.route("/sw.js")
def service_worker():
    response = send_from_directory(BASE_DIR / "static", "sw.js")
    response.headers["Cache-Control"] = "no-cache"
    response.headers["Service-Worker-Allowed"] = "/"
    response.mimetype = "application/javascript"
    return response


@app.context_processor
def inject_globals():
    return {
        "session": session,
        "now": now_dt,
        "format_dt": format_dt,
    }

def init_db_postgres():
    db = get_db()
    cur = db.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS funcionarios (
        id SERIAL PRIMARY KEY,
        nome TEXT,
        cargo TEXT
    );
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS registros (
        id SERIAL PRIMARY KEY,
        funcionario_id INTEGER,
        tipo TEXT,
        data TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """)

    db.commit()


# roda automaticamente quando iniciar
try:
    init_db_postgres()
except Exception as e:
    print("Erro ao criar tabelas:", e)

if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=8080, debug=True)
